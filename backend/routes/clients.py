from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel, Field, field_validator, model_validator
from typing import Optional, List, Dict, Any, Tuple
from datetime import date, datetime, timezone
import logging
import os
import re
import secrets
import uuid
from motor.motor_asyncio import AsyncIOMotorClient
from utils.canonical_id import (
    new_entity_id,
    new_internal_diid,
    normalize_annual_diid,
    validate_annual_diid_format,
)
from utils.garden_labels import (
    LABEL_DEW,
    LABEL_ROOT,
    LABEL_BLOOM,
    LABEL_SEED,
    ORDERED_JOURNEY_LABELS,
    normalize_label,
    is_allowed_manual_label,
    label_filter_variants,
    label_stripe_key,
    iris_anniversary_year_from_client,
    iris_label_for_year,
)
from utils.person_name import normalize_person_name
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent.parent
load_dotenv(ROOT_DIR / '.env')


def _name_initial_segment_for_annual_diid(name: str) -> str:
    """Same rules as ``canonical_id._name_initial_segment`` (four A–Z letters from display name)."""

    parts = (name or "").strip().split()

    def letters(w: str) -> str:
        return "".join(c for c in w.upper() if c.isalpha())

    if len(parts) >= 2:
        a = letters(parts[0])[:2].ljust(2, "X")
        b = letters(parts[-1])[:2].ljust(2, "X")
        return (a + b)[:4]
    if len(parts) == 1:
        p = letters(parts[0])
        return (p + "XXXX")[:4]
    return "XXXX"


def _build_annual_diid_from_name_yymm(name: str, yymm_digits: str) -> Optional[str]:
    """Build full Annual DIID from name + 4 YYMM digits (used when the sheet has YYMM only)."""

    t = (yymm_digits or "").strip()
    m = re.fullmatch(r"(\d{4})", t)
    if not m:
        return None
    t = m.group(1)
    yy, mm = int(t[:2]), int(t[2:4])
    if mm < 1 or mm > 12:
        return None
    initials = _name_initial_segment_for_annual_diid(name or "")
    return f"{initials}{yy:02d}{mm:02d}"


router = APIRouter(prefix="/api/clients", tags=["Clients"])

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

logger = logging.getLogger(__name__)

# How clients first found Divine Iris (CRM attribution)
DISCOVERY_SOURCES: Tuple[str, ...] = (
    "Ads",
    "Instagram",
    "Youtube",
    "Facebook",
    "Google",
    "Website",
    "Referral",
    "Other",
)
_DISCOVERY_LOWER = {s.lower(): s for s in DISCOVERY_SOURCES}


def normalize_discovery_source_value(raw: Optional[str]) -> Optional[str]:
    t = (raw or "").strip()
    if not t:
        return None
    canon = _DISCOVERY_LOWER.get(t.lower())
    if not canon:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid how-they-found-us value. Use one of: {', '.join(DISCOVERY_SOURCES)}",
        )
    return canon


async def resolve_referrer_client(db, self_client_id: str, referrer_id_raw: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    rid = (referrer_id_raw or "").strip()
    if not rid:
        return None, None
    if rid == self_client_id:
        raise HTTPException(status_code=400, detail="Referrer cannot be the same client")
    ref = await db.clients.find_one({"id": rid}, {"_id": 0, "name": 1, "email": 1})
    if not ref:
        raise HTTPException(status_code=404, detail="Referrer UUID not found in Client Garden")
    raw_nm = (ref.get("name") or "").strip()
    if raw_nm:
        display = normalize_person_name(raw_nm)
    else:
        display = (ref.get("email") or "").strip() or rid
    return rid, display


def _first_program_title(conversions: Optional[List]) -> str:
    """Earliest paid conversion by ``date`` — program title they first joined with."""
    convs = [c for c in (conversions or []) if (c.get("program_title") or "").strip()]
    if not convs:
        return ""
    convs_sorted = sorted(convs, key=lambda c: (c.get("date") or ""))
    return (convs_sorted[0].get("program_title") or "").strip()


def effective_first_program(client_doc: dict) -> str:
    """CRM override ``first_program_manual`` wins; else earliest conversion title."""
    manual = (client_doc.get("first_program_manual") or "").strip()
    if manual:
        return manual
    return _first_program_title(client_doc.get("conversions"))


def normalize_email(email: str) -> str:
    return (email or "").strip().lower()


def normalize_phone(phone: str) -> str:
    """Strip separators; preserve leading ``+`` for country codes (E.164-style storage)."""
    t = (phone or "").strip()
    if not t:
        return ""
    t = re.sub(r"[\s\-\(\)\.]", "", t)
    if t.startswith("+"):
        digits = "".join(c for c in t[1:] if c.isdigit())
        return f"+{digits}" if digits else ""
    return "".join(c for c in t if c.isdigit())


def _coerce_client_created_at_iso(raw: str) -> str:
    """Admin CRM: accept YYYY-MM (first of month), YYYY-MM-DD, or ISO 8601; store UTC ISO string."""
    t = (raw or "").strip()
    if not t:
        raise HTTPException(status_code=400, detail="First seen date is empty.")
    if re.fullmatch(r"\d{4}-\d{2}", t):
        try:
            d = datetime.strptime(t + "-01", "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid First seen month.")
        return datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=timezone.utc).isoformat()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", t):
        try:
            d = datetime.strptime(t, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid First seen date.")
        return datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=timezone.utc).isoformat()
    try:
        p = datetime.fromisoformat(t.replace("Z", "+00:00"))
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="First seen must be YYYY-MM, YYYY-MM-DD, or a valid ISO 8601 datetime.",
        )
    if p.tzinfo is None:
        p = p.replace(tzinfo=timezone.utc)
    return p.astimezone(timezone.utc).isoformat()


_DIID_MIDDLE_RE = re.compile(r"^[A-Za-z]{4}\d{4}$")


def _split_internal_diid(diid: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    d = (diid or "").strip()
    if not d or not d.upper().startswith("DIID-"):
        return None, None
    parts = d.split("-")
    if len(parts) < 3:
        return None, None
    middle = parts[1].strip().upper()
    suffix = parts[-1].strip().upper()
    if not middle or not suffix:
        return None, None
    return middle, suffix


def _validate_diid_middle(middle: str) -> str:
    m = (middle or "").strip().upper()
    if not _DIID_MIDDLE_RE.match(m):
        raise HTTPException(
            status_code=400,
            detail="DIID middle must be 4 letters (A–Z) + 4 digits (YYMM), e.g. ABCD2404",
        )
    return m


async def _new_diid_with_middle(db, client_id: str, cl: dict, middle_raw: str) -> str:
    mid = _validate_diid_middle(middle_raw)
    _, suf = _split_internal_diid(cl.get("diid"))
    if not suf:
        suf = secrets.token_hex(4).upper()
    new_diid = f"DIID-{mid}-{suf}"
    dupe = await db.clients.find_one({"diid": new_diid, "id": {"$ne": client_id}})
    if dupe:
        raise HTTPException(status_code=409, detail="That DIID is already in use by another client")
    return new_diid


async def ensure_client_from_enrollment_lead(enrollment: Dict[str, Any]) -> None:
    """
    Create or update a Client Garden row for an enrollment booker. New clients get DID + DIID.
    Re-running for the same enrollment id is a no-op for timeline (uses lead_enrollment_ids).
    """
    try:
        email = normalize_email(enrollment.get("booker_email", ""))
        if not email:
            return
        phone = normalize_phone(
            str(enrollment.get("phone") or enrollment.get("booker_phone") or "")
        )
        eid = str(enrollment.get("id") or "").strip()
        raw_nm = (enrollment.get("booker_name") or "").strip() or (
            email.split("@")[0] if email else ""
        )
        name = normalize_person_name(raw_nm) if raw_nm else ""
        program_title = (enrollment.get("item_title") or "").strip()
        now = datetime.now(timezone.utc).isoformat()
        created_ref = enrollment.get("created_at") or now

        or_query: List[Dict[str, Any]] = [{"email": email}]
        if phone:
            or_query.append({"phone": phone})
        existing = await db.clients.find_one({"$or": or_query})

        detail = " — ".join(p for p in (program_title, eid) if p)[:500] or (eid or "Enrollment")
        timeline_entry = {
            "type": "Enrollment",
            "detail": detail,
            "date": created_ref,
        }

        if existing:
            lead_ids = list(existing.get("lead_enrollment_ids") or [])
            is_new_lead = bool(eid) and eid not in lead_ids
            set_fields: Dict[str, Any] = {"updated_at": now}
            if name and not (existing.get("name") or "").strip():
                set_fields["name"] = name
            if phone and not (existing.get("phone") or "").strip():
                set_fields["phone"] = phone
            if _client_field_empty(existing, "did"):
                set_fields["did"] = f"DID-{uuid.uuid4().hex[:8].upper()}"
            if _client_field_empty(existing, "diid"):
                anchor = existing.get("created_at") or created_ref
                set_fields["diid"] = new_internal_diid(
                    (existing.get("name") or name or "")[:200], anchor
                )
            backfill_only = {k for k in set_fields.keys() if k != "updated_at"}

            if is_new_lead:
                add_to_set: Dict[str, Any] = {"sources": "Enrollment"}
                if eid:
                    add_to_set["lead_enrollment_ids"] = eid
                await db.clients.update_one(
                    {"id": existing["id"]},
                    {
                        "$addToSet": add_to_set,
                        "$push": {"timeline": timeline_entry},
                        "$set": set_fields,
                    },
                )
            elif backfill_only:
                await db.clients.update_one(
                    {"id": existing["id"]},
                    {"$set": set_fields},
                )
            return

        did = f"DID-{uuid.uuid4().hex[:8].upper()}"
        client_doc: Dict[str, Any] = {
            "id": new_entity_id(),
            "did": did,
            "diid": new_internal_diid(name, created_ref),
            "email": email,
            "phone": phone or "",
            "name": name,
            "label": LABEL_DEW,
            "label_manual": "",
            "sources": ["Enrollment"],
            "conversions": [],
            "timeline": [timeline_entry],
            "notes": "",
            "created_at": now,
            "updated_at": now,
        }
        if eid:
            client_doc["lead_enrollment_ids"] = [eid]
        await db.clients.insert_one(client_doc)
    except Exception as e:
        logger.warning("ensure_client_from_enrollment_lead failed: %s", e)


def enrollment_counts_as_paid_conversion(enrollment: Dict[str, Any], txs_for_enrollment: List[dict]) -> bool:
    """
    A conversion is recorded only after the payment/checkout loop succeeds.
    Enrollments move to status 'completed' when Stripe (or webhook), free checkout,
    or India manual proof flow finishes; we also require a paid transaction when one exists.
    """
    if (enrollment.get("status") or "").strip().lower() != "completed":
        return False
    if txs_for_enrollment:
        return any(str(t.get("payment_status", "")).strip().lower() == "paid" for t in txs_for_enrollment)
    # Legacy completed rows with no transaction document — still count if marked paid/free on enrollment
    return bool(enrollment.get("paid_at")) or bool(enrollment.get("is_free"))


async def compute_label(client_doc: dict) -> str:
    """
    Auto-label from paid conversions only (see sync). Manual garden label always wins.

    - No manual override and zero paid-program conversions → Dew (leads, first-time,
      or checkout not finished).
    - With conversions, stage advances (Seed → Root → Bloom / Iris years) from program data.
    """
    lm = (client_doc.get("label_manual") or "").strip()
    if lm:
        return normalize_label(lm)

    conversions = client_doc.get("conversions", [])
    if not conversions:
        return LABEL_DEW

    flagship_count = sum(1 for c in conversions if c.get("is_flagship"))
    workshop_count = sum(1 for c in conversions if not c.get("is_flagship"))
    total = len(conversions)

    # Annual / year-long enrollment → Iris year 1–12 from subscription start when present
    has_annual = any(
        "annual" in (c.get("tier_label", "") or "").lower()
        or (c.get("duration_unit", "") == "year")
        for c in conversions
    )
    if has_annual:
        y = iris_anniversary_year_from_client(client_doc)
        return iris_label_for_year(y)
    if total >= 3 or (flagship_count >= 1 and workshop_count >= 1):
        return LABEL_BLOOM
    if flagship_count >= 1:
        return LABEL_ROOT
    if workshop_count >= 1 or total >= 1:
        return LABEL_SEED
    return LABEL_DEW


def _client_field_empty(doc: dict, key: str) -> bool:
    v = doc.get(key)
    return v is None or (isinstance(v, str) and not str(v).strip())


async def backfill_missing_client_identifiers() -> int:
    """
    Fill `did` (DID-xxxxxxxx) and branded `diid` (DIID-INITIALSyyMM-HEX) on legacy rows.
    Does not change `updated_at` so Client Garden sort order stays stable.
    """
    n = 0
    async for cl in db.clients.find(
        {},
        {"_id": 1, "name": 1, "created_at": 1, "updated_at": 1, "did": 1, "diid": 1},
    ):
        set_doc: Dict[str, Any] = {}
        if _client_field_empty(cl, "did"):
            set_doc["did"] = f"DID-{uuid.uuid4().hex[:8].upper()}"
        anchor = cl.get("created_at") or cl.get("updated_at") or datetime.now(timezone.utc).isoformat()
        if _client_field_empty(cl, "diid"):
            set_doc["diid"] = new_internal_diid(cl.get("name") or "", anchor)
        if set_doc:
            await db.clients.update_one({"_id": cl["_id"]}, {"$set": set_doc})
            n += 1
    return n


# ========== SYNC / BACKFILL ==========

@router.post("/sync")
async def sync_clients():
    """Backfill/sync client data from contacts, interests, questions; enrollments count as conversions only after paid checkout."""
    stats = {
        "new_clients": 0,
        "updated": 0,
        "total_sources_scanned": 0,
        "conversions_pruned_clients": 0,
        "identifiers_backfilled": 0,
    }

    # Helper to upsert a client
    async def upsert_client(email: str, phone: str = "", name: str = "", source: str = "", source_detail: str = "", source_date: str = ""):
        email = normalize_email(email)
        phone = normalize_phone(phone)
        name_raw = (name or "").strip()
        name = normalize_person_name(name_raw) if name_raw else ""
        if not email and not phone:
            return

        # Find existing by email or phone
        query = []
        if email:
            query.append({"email": email})
        if phone:
            query.append({"phone": phone})
        existing = await db.clients.find_one({"$or": query}) if query else None

        now = datetime.now(timezone.utc).isoformat()
        timeline_entry = {
            "type": source,
            "detail": source_detail,
            "date": source_date or now,
        }

        if existing:
            update = {"$addToSet": {"timeline": timeline_entry, "sources": source}}
            if name and not existing.get("name"):
                update["$set"] = {"name": name}
            if phone and not existing.get("phone"):
                update.setdefault("$set", {})["phone"] = phone
            await db.clients.update_one({"id": existing["id"]}, update)
            stats["updated"] += 1
        else:
            # Generate DID for new clients (first-time joiners)
            did = f"DID-{str(uuid.uuid4())[:8].upper()}"
            client_doc = {
                "id": new_entity_id(),
                "did": did,
                "diid": new_internal_diid(name or "", now),
                "email": email,
                "phone": phone,
                "name": name,
                "label": LABEL_DEW,
                "label_manual": "",
                "sources": [source],
                "conversions": [],
                "timeline": [timeline_entry],
                "notes": "",
                "created_at": now,
                "updated_at": now,
            }
            await db.clients.insert_one(client_doc)
            stats["new_clients"] += 1

    # 1. Contact form submissions
    contacts = await db.quote_requests.find({}, {"_id": 0}).to_list(2000)
    for c in contacts:
        await upsert_client(
            email=c.get("email", ""),
            phone=c.get("phone", ""),
            name=c.get("name", ""),
            source="Contact Form",
            source_detail=c.get("message", "")[:100],
            source_date=c.get("created_at", ""),
        )
    stats["total_sources_scanned"] += len(contacts)

    # 2. Express your interest
    interests = await db.notify_me.find({}, {"_id": 0}).to_list(2000)
    for i in interests:
        await upsert_client(
            email=i.get("email", ""),
            name=i.get("name", ""),
            source="Express Interest",
            source_detail=i.get("program_title", ""),
            source_date=i.get("created_at", ""),
        )
    stats["total_sources_scanned"] += len(interests)

    # 3. Questions
    questions = await db.session_questions.find({}, {"_id": 0}).to_list(2000)
    for q in questions:
        await upsert_client(
            email=q.get("email", ""),
            name=q.get("name", ""),
            source="Question",
            source_detail=q.get("question", "")[:100],
            source_date=q.get("created_at", ""),
        )
    stats["total_sources_scanned"] += len(questions)

    # 4. Enrollments + Payment Transactions (the conversion data)
    # First build a map of stripe_session_id -> transaction for item_title lookup
    transactions = await db.payment_transactions.find({}, {"_id": 0}).to_list(5000)
    tx_by_enrollment = {}
    for tx in transactions:
        eid = tx.get("enrollment_id", "")
        if eid:
            tx_by_enrollment.setdefault(eid, []).append(tx)

    enrollments = await db.enrollments.find({}, {"_id": 0}).to_list(2000)
    programs_cache = {}
    all_programs = await db.programs.find({}, {"_id": 0}).to_list(100)
    for p in all_programs:
        programs_cache[p["id"]] = p

    sessions_cache = {}
    all_sessions = await db.sessions.find({"token": {"$exists": False}}, {"_id": 0}).to_list(100)
    for ss in all_sessions:
        sessions_cache[ss["id"]] = ss

    paid_conversion_enrollment_ids = {
        e["id"]
        for e in enrollments
        if enrollment_counts_as_paid_conversion(e, tx_by_enrollment.get(e.get("id", ""), []))
    }

    for e in enrollments:
        email = normalize_email(e.get("booker_email", ""))
        phone = normalize_phone(e.get("booker_phone", ""))
        if not email:
            continue

        enrollment_txs = tx_by_enrollment.get(e.get("id", ""), [])
        if not enrollment_counts_as_paid_conversion(e, enrollment_txs):
            continue

        program_id = e.get("program_id", "") or e.get("selected_program_id", "")
        session_id = e.get("session_id", "") or e.get("selected_session_id", "")
        program = programs_cache.get(program_id, {})

        # Try to get title from transaction data first
        program_title = ""
        if enrollment_txs:
            program_title = enrollment_txs[0].get("item_title", "")
            if not program_id:
                program_id = enrollment_txs[0].get("item_id", "")
                program = programs_cache.get(program_id, {})

        if not program_title:
            program_title = e.get("program_title", "") or e.get("selected_program_title", "") or program.get("title", "")
        if not program_title and session_id:
            program_title = sessions_cache.get(session_id, {}).get("title", "")

        is_flagship = program.get("is_flagship", False)
        status = e.get("status", "") or "completed"

        # Find or create client
        query = [{"email": email}]
        if phone:
            query.append({"phone": phone})
        existing = await db.clients.find_one({"$or": query})

        conversion_entry = {
            "enrollment_id": e.get("id", ""),
            "program_id": program_id,
            "program_title": program_title,
            "is_flagship": is_flagship,
            "status": status,
            "item_type": e.get("item_type", ""),
            "tier_label": e.get("tier_label", ""),
            "duration_unit": e.get("duration_unit", ""),
            "date": e.get("paid_at") or e.get("updated_at") or e.get("created_at", ""),
        }
        timeline_entry = {
            "type": "Enrollment (paid)",
            "detail": f"{program_title} — payment complete",
            "date": e.get("paid_at") or e.get("updated_at") or e.get("created_at", ""),
        }

        if existing:
            # Check if this enrollment already tracked
            existing_enrollment_ids = [c.get("enrollment_id") for c in existing.get("conversions", [])]
            if e.get("id") not in existing_enrollment_ids:
                await db.clients.update_one(
                    {"id": existing["id"]},
                    {
                        "$push": {"conversions": conversion_entry, "timeline": timeline_entry},
                        "$addToSet": {"sources": "Enrollment"},
                    }
                )
            # Update name/phone if missing
            update_set = {}
            booker_nm = normalize_person_name((e.get("booker_name") or "").strip())
            if booker_nm and not existing.get("name"):
                update_set["name"] = booker_nm
            if phone and not existing.get("phone"):
                update_set["phone"] = phone
            if update_set:
                await db.clients.update_one({"id": existing["id"]}, {"$set": update_set})
        else:
            did = f"DID-{str(uuid.uuid4())[:8].upper()}"
            now = datetime.now(timezone.utc).isoformat()
            booker_nm = normalize_person_name((e.get("booker_name") or "").strip())
            client_doc = {
                "id": new_entity_id(),
                "did": did,
                "diid": new_internal_diid(booker_nm or "", now),
                "email": email,
                "phone": phone,
                "name": booker_nm,
                "label": LABEL_DEW,
                "label_manual": "",
                "sources": ["Enrollment"],
                "conversions": [conversion_entry],
                "timeline": [timeline_entry],
                "notes": "",
                "created_at": now,
                "updated_at": now,
            }
            await db.clients.insert_one(client_doc)
            stats["new_clients"] += 1

    stats["total_sources_scanned"] += len(enrollments)

    # 4b Remove conversion rows for enrollments that never finished payment
    all_for_prune = await db.clients.find({}, {"_id": 0, "id": 1, "conversions": 1}).to_list(5000)
    for cl in all_for_prune:
        convs = cl.get("conversions") or []
        if not convs:
            continue

        filtered = [
            c
            for c in convs
            if (not c.get("enrollment_id")) or (c.get("enrollment_id") in paid_conversion_enrollment_ids)
        ]
        if len(filtered) != len(convs):
            await db.clients.update_one(
                {"id": cl["id"]},
                {"$set": {"conversions": filtered, "updated_at": datetime.now(timezone.utc).isoformat()}},
            )
            stats["conversions_pruned_clients"] += 1

    # 5. Normalize legacy short labels and recompute automatic labels
    all_clients = await db.clients.find({}, {"_id": 0}).to_list(5000)
    now_iso = datetime.now(timezone.utc).isoformat()
    for cl in all_clients:
        lm = (cl.get("label_manual") or "").strip()
        set_doc: Dict[str, Any] = {}
        if lm:
            nlm = normalize_label(lm)
            if nlm != lm:
                set_doc["label_manual"] = nlm
                cl = {**cl, "label_manual": nlm}
        new_label = await compute_label(cl)
        if new_label != cl.get("label"):
            set_doc["label"] = new_label
        if set_doc:
            set_doc["updated_at"] = now_iso
            await db.clients.update_one({"id": cl["id"]}, {"$set": set_doc})
            stats["updated"] += 1

    stats["identifiers_backfilled"] = await backfill_missing_client_identifiers()

    return {"message": "Sync complete", "stats": stats}


# ========== LIST / SEARCH ==========

@router.get("")
async def list_clients(label: Optional[str] = None, search: Optional[str] = None, intake_pending: Optional[str] = None):
    query = {}
    if label:
        variants = label_filter_variants(label)
        if len(variants) == 1:
            query["label"] = variants[0]
        else:
            query["label"] = {"$in": variants}
    if intake_pending == "true":
        query["intake_pending"] = True
        query["portal_login_allowed"] = False
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        query["$or"] = [
            {"name": search_regex},
            {"email": search_regex},
            {"phone": search_regex},
            {"household_key": search_regex},
            {"did": search_regex},
            {"diid": search_regex},
            {"id": search_regex},
        ]
    # Stable order so inline edits + refetch do not jump the row to the top (``updated_at`` changes every save).
    clients_list = await db.clients.find(query, {"_id": 0}).sort(
        [("created_at", 1), ("name", 1), ("id", 1)]
    ).to_list(1000)
    for cl in clients_list:
        cl["first_program"] = effective_first_program(cl)
    return clients_list


class ClientManualCreate(BaseModel):
    """Create a single client from Client Garden (trial / manual entry). Name only is allowed."""
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    notes: Optional[str] = None
    label_manual: Optional[str] = None  # empty = auto Dew


@router.post("")
async def create_client_manual(data: ClientManualCreate):
    """Manually add a client. Requires name; email and phone are optional."""
    name = normalize_person_name((data.name or "").strip())
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")

    email_n = normalize_email(data.email or "")
    phone_n = normalize_phone(data.phone or "")

    label_manual = (data.label_manual or "").strip()
    if label_manual:
        if not is_allowed_manual_label(label_manual):
            raise HTTPException(
                status_code=400,
                detail="Invalid garden label — use Dew / Seed / Root / Bloom, Iris — The Seeker, Iris year 1–12, Purple Bees, or Iris Bees (short or full title).",
            )
        label_manual = normalize_label(label_manual)

    if phone_n:
        existing_phone = await db.clients.find_one({"phone": phone_n})
        if existing_phone:
            raise HTTPException(
                status_code=409,
                detail="A client with this phone number already exists",
            )

    now = datetime.now(timezone.utc).isoformat()
    did = f"DID-{str(uuid.uuid4())[:8].upper()}"
    cid = new_entity_id()
    initial_label = label_manual if label_manual else LABEL_DEW
    timeline_entry = {
        "type": "Manual",
        "detail": "Added from Client Garden",
        "date": now,
    }
    client_doc = {
        "id": cid,
        "did": did,
        "diid": new_internal_diid(name, now),
        "email": email_n,
        "phone": phone_n,
        "name": name,
        "label": initial_label,
        "label_manual": label_manual,
        "sources": ["Manual"],
        "conversions": [],
        "timeline": [timeline_entry],
        "notes": (data.notes or "").strip(),
        "created_at": now,
        "updated_at": now,
        "portal_login_allowed": True,
    }
    await db.clients.insert_one(client_doc)

    if not label_manual:
        new_label = await compute_label(client_doc)
        if new_label != initial_label:
            await db.clients.update_one({"id": cid}, {"$set": {"label": new_label, "updated_at": now}})

    return {"message": "Client created", "id": cid}


@router.get("/stats")
async def client_stats():
    pipeline = [
        {"$group": {"_id": "$label", "count": {"$sum": 1}}},
    ]
    results = await db.clients.aggregate(pipeline).to_list(200)
    label_counts = {r["_id"]: r["count"] for r in results}
    total = sum(label_counts.values())
    return {"total": total, "by_label": label_counts}


@router.get("/garden-label-options")
async def garden_label_options():
    """Canonical Client Garden labels for admin dropdowns (order matches journey)."""
    return {"labels": ORDERED_JOURNEY_LABELS}


@router.get("/discovery-options")
async def discovery_options():
    """Allowed values for how the client first found Divine Iris."""
    return {"sources": list(DISCOVERY_SOURCES)}


@router.get("/annual-portal-subscribers")
async def list_annual_portal_subscribers():
    """
    Clients with ``annual_member_dashboard`` True — same cohort as **Annual program = Yes** in the
    main Client Garden grid. Includes rows with Google login blocked; the UI can flag those.
    (Excel upload still skips ``portal_login_allowed`` False until sign-in is allowed.)

    Overdue ``annual_subscription.end_date`` is **not** bulk-cleared here — that used to wipe
    ``annual_member_dashboard`` as soon as admins opened this tab, making saves look broken.
    Student-facing routes still call :func:`persist_annual_member_expiry_if_overdue` so effective
    portal access stays correct.
    """
    query = {"annual_member_dashboard": True}
    proj = {
        "_id": 0,
        "id": 1,
        "name": 1,
        "email": 1,
        "household_key": 1,
        "is_primary_household_contact": 1,
        "annual_subscription": 1,
        "portal_login_allowed": 1,
    }
    rows = (
        await db.clients.find(query, proj).sort([("name", 1)]).to_list(5000)
    )
    return {"clients": rows}


HOME_COMING_SKU = "home_coming"

# Sacred Home nudge when annual_subscription.end_date is within this many days (inclusive).
_ANNUAL_RENEWAL_WARN_DAYS = 30


def parse_annual_subscription_end_date(client: Dict[str, Any]) -> Optional[date]:
    """Calendar end date from Client Garden ``annual_subscription`` (YYYY-MM-DD), or None."""
    sub = client.get("annual_subscription") or {}
    end = (sub.get("end_date") or "").strip()
    if not end or len(end) < 10:
        return None
    try:
        return datetime.strptime(end[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def utc_today() -> date:
    return datetime.now(timezone.utc).date()


def annual_subscription_period_expired(client: Dict[str, Any]) -> bool:
    """True when ``annual_subscription.end_date`` is strictly before today (UTC calendar date)."""
    end_d = parse_annual_subscription_end_date(client)
    if not end_d:
        return False
    return utc_today() > end_d


async def persist_annual_member_expiry_if_overdue(db, client: Dict[str, Any]) -> Dict[str, Any]:
    """If CRM annual flag is on but Home Coming end_date has passed, set ``annual_member_dashboard`` False."""
    if not isinstance(client, dict) or not client.get("id"):
        return client
    if not bool(client.get("annual_member_dashboard")):
        return client
    if not annual_subscription_period_expired(client):
        return client
    now = datetime.now(timezone.utc).isoformat()
    await db.clients.update_one(
        {"id": client["id"]},
        {"$set": {"annual_member_dashboard": False, "updated_at": now}},
    )
    out = dict(client)
    out["annual_member_dashboard"] = False
    return out


async def expire_all_overdue_annual_dashboard_clients(db) -> int:
    """Clear ``annual_member_dashboard`` for every client whose ``annual_subscription.end_date`` has passed.

    Not called from the admin list/export endpoints (that caused saves to appear to vanish when
    staff opened Client Garden). Student routes still use :func:`persist_annual_member_expiry_if_overdue`.
    """
    q = {
        "annual_member_dashboard": True,
        "annual_subscription.end_date": {"$regex": r"^\d{4}-\d{2}-\d{2}"},
    }
    stale_ids: List[str] = []
    async for doc in db.clients.find(q, {"_id": 0, "id": 1, "annual_subscription": 1}):
        if annual_subscription_period_expired(doc):
            cid = doc.get("id")
            if cid:
                stale_ids.append(cid)
    if not stale_ids:
        return 0
    now = datetime.now(timezone.utc).isoformat()
    res = await db.clients.update_many(
        {"id": {"$in": stale_ids}},
        {"$set": {"annual_member_dashboard": False, "updated_at": now}},
    )
    return int(res.modified_count)


def annual_renewal_reminder_for_portal(client: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Banner payload for Sacred Home when ``annual_subscription.end_date`` is set:
    expiring within ``_ANNUAL_RENEWAL_WARN_DAYS`` (and still annual), or already ended.
    """
    end_d = parse_annual_subscription_end_date(client)
    if not end_d:
        return None
    today = utc_today()
    days_left = (end_d - today).days
    end_s = end_d.isoformat()
    if days_left < 0:
        return {
            "kind": "expired",
            "end_date": end_s,
            "message": (
                "Your annual Sacred Home period ended on this date. Renew with Divine Iris to keep "
                "member pricing, package inclusions, and your journey benefits."
            ),
        }
    if days_left <= _ANNUAL_RENEWAL_WARN_DAYS and bool(client.get("annual_member_dashboard")):
        if days_left == 0:
            msg = (
                "Your annual plan ends today. Renew soon to continue uninterrupted member pricing and benefits."
            )
        elif days_left == 1:
            msg = "Your annual plan ends tomorrow. Renew to keep your member benefits and growth path."
        else:
            msg = (
                f"Your annual plan ends in {days_left} days ({end_s}). Renew on time to keep member pricing and Sacred Home benefits."
            )
        return {
            "kind": "expiring",
            "end_date": end_s,
            "days_remaining": days_left,
            "message": msg,
        }
    return None


class AnnualSubscriptionUsagePatch(BaseModel):
    awrp_months_used: Optional[int] = None
    mmm_months_used: Optional[int] = None
    turbo_sessions_used: Optional[int] = None
    meta_downloads_used: Optional[int] = None

    @field_validator(
        "awrp_months_used",
        "mmm_months_used",
        "turbo_sessions_used",
        "meta_downloads_used",
    )
    @classmethod
    def _non_negative(cls, v: Optional[int]) -> Optional[int]:
        if v is None:
            return v
        if int(v) < 0:
            raise ValueError("usage counts must be non-negative")
        return int(v)


class AnnualSubscriptionUpdate(BaseModel):
    """
    Partial update for annual_subscription on a client (Home Coming SKU, DIID, dates, usage).
    Omitted fields are unchanged. Explicit null clears start_date, end_date, annual_diid, etc.
    """

    annual_diid: Optional[str] = None
    package_sku: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    usage: Optional[AnnualSubscriptionUsagePatch] = None
    """manual = admin/backfill; system = derived from bookings (future)."""

    usage_source: Optional[str] = None

    @field_validator("usage_source")
    @classmethod
    def _usage_source(cls, v: Optional[str]) -> Optional[str]:
        if v is None:
            return v
        t = (v or "").strip().lower()
        if t not in ("manual", "system"):
            raise ValueError("usage_source must be 'manual' or 'system'")
        return t


def _coerce_annual_date_field(raw: Optional[str], field_label: str) -> Any:
    """
    For PATCH: None means caller did not include field. '' or explicit null (becomes None in body)
    clears. Non-empty must be YYYY-MM-DD.
    Returns NOT_PROVIDED sentinel vs '' vs 'YYYY-MM-DD' — actually we use exclude_unset on model.
    """
    if raw is None:
        return None
    s = (raw or "").strip()
    if not s:
        return ""
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        try:
            datetime.strptime(s[:10], "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid {field_label} — use YYYY-MM-DD",
            )
        return s[:10]
    raise HTTPException(
        status_code=400,
        detail=f"Invalid {field_label} — use YYYY-MM-DD",
    )


def _merge_annual_subscription(
    prev: Optional[Dict[str, Any]], patch: Dict[str, Any]
) -> Dict[str, Any]:
    out: Dict[str, Any] = dict(prev or {})

    if "annual_diid" in patch:
        v = patch["annual_diid"]
        if v is None or v == "":
            out.pop("annual_diid", None)
        else:
            out["annual_diid"] = v

    if "package_sku" in patch:
        v = patch["package_sku"]
        if v is None or v == "":
            out.pop("package_sku", None)
        else:
            out["package_sku"] = v

    if "start_date" in patch:
        v = patch["start_date"]
        if v is None or v == "":
            out.pop("start_date", None)
        else:
            out["start_date"] = v

    if "end_date" in patch:
        v = patch["end_date"]
        if v is None or v == "":
            out.pop("end_date", None)
        else:
            out["end_date"] = v

    if "usage_source" in patch:
        v = patch["usage_source"]
        if v is None or v == "":
            out.pop("usage_source", None)
        else:
            out["usage_source"] = v

    if "usage" in patch:
        u = patch["usage"]
        if u is None:
            out.pop("usage", None)
        elif isinstance(u, dict):
            u_prev: Dict[str, Any] = dict(out.get("usage") or {})
            for uk, uv in u.items():
                if uv is None:
                    u_prev.pop(uk, None)
                else:
                    u_prev[uk] = int(uv)
            out["usage"] = u_prev

    out.pop("awrp_year_label", None)
    return out


@router.patch("/{client_id}/annual-subscription")
async def patch_annual_subscription(client_id: str, data: AnnualSubscriptionUpdate):
    """
    Update Home Coming / annual subscription fields. annual_diid must be unique (FFLLYYMM).
    """
    cl = await db.clients.find_one({"id": client_id})
    if not cl:
        raise HTTPException(status_code=404, detail="Client not found")

    raw = data.model_dump(exclude_unset=True)
    patch: Dict[str, Any] = {}

    if "annual_diid" in raw:
        ad_raw = raw["annual_diid"]
        if ad_raw is None or (isinstance(ad_raw, str) and not ad_raw.strip()):
            patch["annual_diid"] = None
        else:
            ad = normalize_annual_diid(str(ad_raw))
            if not validate_annual_diid_format(ad):
                raise HTTPException(
                    status_code=400,
                    detail="annual_diid must be 4 letters + YYMM (e.g. ANRA2504)",
                )
            dup = await db.clients.find_one(
                {
                    "annual_subscription.annual_diid": ad,
                    "id": {"$ne": client_id},
                }
            )
            if dup:
                raise HTTPException(
                    status_code=400,
                    detail="This annual DIID is already assigned to another client",
                )
            patch["annual_diid"] = ad

    if "package_sku" in raw:
        sku_raw = raw["package_sku"]
        if sku_raw is None or (isinstance(sku_raw, str) and not str(sku_raw).strip()):
            patch["package_sku"] = None
        else:
            sku = str(sku_raw).strip().lower()
            if sku != HOME_COMING_SKU:
                raise HTTPException(
                    status_code=400,
                    detail=f"package_sku must be '{HOME_COMING_SKU}'",
                )
            patch["package_sku"] = sku

    if "start_date" in raw:
        patch["start_date"] = _coerce_annual_date_field(
            raw["start_date"], "start_date"
        )

    if "end_date" in raw:
        patch["end_date"] = _coerce_annual_date_field(raw["end_date"], "end_date")

    if "usage_source" in raw:
        patch["usage_source"] = raw["usage_source"]

    if "usage" in raw:
        if raw["usage"] is None:
            patch["usage"] = None
        else:
            patch["usage"] = raw["usage"]

    prev_sub: Dict[str, Any] = dict(cl.get("annual_subscription") or {})
    merged = _merge_annual_subscription(prev_sub, patch)

    now = datetime.now(timezone.utc).isoformat()
    await db.clients.update_one(
        {"id": client_id},
        {"$set": {"annual_subscription": merged, "updated_at": now}},
    )
    return {"annual_subscription": merged}


def _annual_upload_norm_header(h: Any) -> str:
    s = (str(h or "").strip().lower().replace("*", "").replace("\n", " "))
    return re.sub(r"\s+", " ", s)


def _annual_upload_cell_str(row: tuple, idx: Optional[int]) -> str:
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _annual_diid_cell_raw(row: tuple, idx: Optional[int]) -> str:
    """Excel may store YYMM as int/float (e.g. 2503); normalize to digits without .0."""
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        f = float(v)
        if f == int(f) and -1e12 <= f <= 1e12:
            return str(int(f))
    s = str(v).strip()
    if re.match(r"^\d+\.0+$", s):
        s = s.split(".", 1)[0]
    return s


def _display_name_for_portal_upload(
    row: tuple, col_map: Dict[str, int], cl: Dict[str, Any]
) -> str:
    if "row_name" in col_map:
        n = _annual_upload_cell_str(row, col_map["row_name"]).strip()
        if n:
            return normalize_person_name(n)
    cur = (cl.get("name") or "").strip()
    return normalize_person_name(cur) if cur else ""


async def _find_single_client_by_name(name: str) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    n = (name or "").strip()
    if not n:
        return None, None
    esc = re.escape(n)
    cur = db.clients.find({"name": {"$regex": f"^{esc}$", "$options": "i"}}).limit(2)
    matches: List[Dict[str, Any]] = []
    async for doc in cur:
        matches.append(doc)
    if not matches:
        return None, f"no client with exact name {n!r} — add Client id, use a unique Name + HOUSEHOLD, or leave Email if present"
    if len(matches) > 1:
        return None, f"more than one client with name {n!r} — set Client id or disambiguate with HOUSEHOLD on the row"
    return matches[0], None


async def _find_client_by_name_and_household(
    name: str, household_key: str
) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    n = (name or "").strip()
    hk = (household_key or "").strip()
    if not n or not hk:
        return None, None
    esc_n = re.escape(n)
    cur = db.clients.find(
        {
            "name": {"$regex": f"^{esc_n}$", "$options": "i"},
            "household_key": hk,
        }
    ).limit(2)
    matches: List[Dict[str, Any]] = []
    async for doc in cur:
        matches.append(doc)
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        return None, f"more than one client with name {n!r} in household {hk!r} — set Client id"
    return None, None


async def _create_client_for_annual_portal_upload(
    row: tuple,
    col_map: Dict[str, int],
    row_idx: int,
    name: str,
    errors: List[str],
) -> Optional[Dict[str, Any]]:
    """New Client Garden row with generated id; annual_member_dashboard on. Email optional."""
    name = normalize_person_name((name or "").strip())
    if not name:
        return None
    email_n = normalize_email(_annual_upload_cell_str(row, col_map.get("email")))
    if email_n:
        existing = await db.clients.find_one({"email": email_n})
        if existing:
            errors.append(
                f"Row {row_idx}: email {email_n} is already used by another client — use that row's Client id, or clear Email to create a member without login email"
            )
            return None
    now = datetime.now(timezone.utc).isoformat()
    cid = new_entity_id()
    did = f"DID-{str(uuid.uuid4())[:8].upper()}"
    timeline_entry = {
        "type": "Manual",
        "detail": "Created from annual portal Excel import",
        "date": now,
    }
    doc: Dict[str, Any] = {
        "id": cid,
        "did": did,
        "diid": new_internal_diid(name, now),
        "email": email_n,
        "phone": "",
        "name": name,
        "label": iris_label_for_year(1),
        "label_manual": iris_label_for_year(1),
        "sources": ["Annual portal import"],
        "conversions": [],
        "timeline": [timeline_entry],
        "notes": "",
        "created_at": now,
        "updated_at": now,
        "portal_login_allowed": True,
        "annual_member_dashboard": True,
    }
    if "household_key" in col_map:
        hk_raw = _annual_upload_cell_str(row, col_map["household_key"]).strip()
        if hk_raw:
            doc["household_key"] = hk_raw[:200]
    if "is_primary_household_contact" in col_map:
        prim_raw = _annual_upload_cell_str(row, col_map["is_primary_household_contact"])
        pt = prim_raw.strip().upper()
        if not pt or pt in ("N", "NO", "0", "FALSE", "-", "—"):
            doc["is_primary_household_contact"] = False
        elif pt in ("Y", "YES", "1", "TRUE"):
            doc["is_primary_household_contact"] = True
        else:
            pl = prim_raw.strip().lower()
            if pl in ("y", "yes", "1", "true"):
                doc["is_primary_household_contact"] = True
            elif pl in ("n", "no", "0", "false"):
                doc["is_primary_household_contact"] = False
            else:
                errors.append(
                    f"Row {row_idx}: PRIMARY must be Y or N (got {prim_raw!r})"
                )
                return None
    await db.clients.insert_one(doc)
    return doc


ANNUAL_PORTAL_UPLOAD_SPECS: List[Tuple[str, Tuple[str, ...]]] = [
    (
        "client_id",
        ("client id", "client_id", "clientid", "uuid", "garden id", "client uuid"),
    ),
    (
        "email",
        (
            "email",
            "e-mail",
            "email id",
            "emailid",
            "e mail",
            "email address",
        ),
    ),
    (
        "row_name",
        ("name", "member name", "client name", "full name", "display name", "client"),
    ),
    ("annual_diid", ("annual diid", "annual_diid", "member diid", "diid")),
    (
        "start_date",
        ("start", "start date", "subscription start", "start_date", "startdate"),
    ),
    (
        "end_date",
        ("end", "end date", "subscription end", "end_date", "enddate"),
    ),
    (
        "package_sku",
        ("package", "package_sku", "sku", "homecoming", "home coming"),
    ),
    (
        "household_key",
        ("household", "household_key", "household key"),
    ),
    (
        "is_primary_household_contact",
        ("primary", "is_primary", "primary household", "is primary"),
    ),
    ("awrp_months_used", ("awrp months used", "awrp_months_used")),
    ("mmm_months_used", ("mmm months used", "mmm_months_used")),
    ("turbo_sessions_used", ("turbo sessions used", "turbo_sessions_used", "turbo")),
    ("meta_downloads_used", ("meta downloads used", "meta_downloads_used", "meta")),
    ("usage_source", ("usage source", "usage_source")),
]


def _annual_portal_upload_col_map(header_cells: List[Any]) -> Dict[str, int]:
    norm = [_annual_upload_norm_header(c) for c in header_cells]
    col_map: Dict[str, int] = {}
    for field, aliases in ANNUAL_PORTAL_UPLOAD_SPECS:
        for i, n in enumerate(norm):
            if n in aliases:
                col_map[field] = i
                break
    return col_map


def _find_annual_portal_header_row(ws: Any) -> Tuple[int, List[Any]]:
    """First row (1-based) that looks like a header: Email, Client id, and/or Name."""
    max_scan = min(30, max(ws.max_row or 1, 1))
    max_c = max(ws.max_column or 40, 1)
    for r in range(1, max_scan + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
        if not any(x not in (None, "") for x in cells):
            continue
        cmap = _annual_portal_upload_col_map(cells)
        if "email" in cmap or "client_id" in cmap or "row_name" in cmap:
            return r, cells
    cells0 = [ws.cell(row=1, column=c).value for c in range(1, max_c + 1)]
    return 1, cells0


def _coerce_upload_date_value(v: Any, label: str) -> Tuple[Optional[str], Optional[str]]:
    """Accept ISO string, Excel date cell, or serial number."""
    if v is None:
        return None, None
    if isinstance(v, datetime):
        return v.date().isoformat(), None
    if isinstance(v, date):
        return v.isoformat(), None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            from openpyxl.utils.datetime import from_excel

            dt = from_excel(float(v))
            return dt.date().isoformat(), None
        except Exception:
            pass
    s = str(v).strip()
    if not s:
        return None, None
    return _coerce_upload_date(s, label)


def _parse_package_cell_for_upload(val: str) -> Optional[str]:
    t = (val or "").strip().lower().replace(" ", "_")
    if not t:
        return None
    if t == HOME_COMING_SKU or t in ("homecoming", "hc"):
        return HOME_COMING_SKU
    if "home" in t and "coming" in t.replace("_", ""):
        return HOME_COMING_SKU
    return "__invalid__"


def _parse_int_cell_strict(val: str) -> Tuple[Optional[int], Optional[str]]:
    s = (val or "").strip()
    if not s:
        return None, None
    try:
        n = int(float(s))
        if n < 0:
            return None, "must be non-negative"
        return n, None
    except ValueError:
        return None, "not a number"


def _coerce_upload_date(s: str, label: str) -> Tuple[Optional[str], Optional[str]]:
    if not (s or "").strip():
        return None, None
    t = str(s).strip()
    if len(t) >= 10 and t[4] == "-" and t[7] == "-":
        try:
            datetime.strptime(t[:10], "%Y-%m-%d")
            return t[:10], None
        except ValueError:
            return None, f"invalid {label} (use YYYY-MM-DD)"
    return None, f"invalid {label} (use YYYY-MM-DD)"


ANNUAL_PORTAL_UPLOAD_PAYLOAD_KEYS = frozenset(
    {
        "household_key",
        "is_primary_household_contact",
        "annual_diid",
        "start_date",
        "end_date",
        "package_sku",
        "usage_source",
        "awrp_months_used",
        "mmm_months_used",
        "turbo_sessions_used",
        "meta_downloads_used",
    }
)

ANNUAL_PORTAL_USAGE_SPECS = (
    ("awrp_months_used", "awrp_months_used"),
    ("mmm_months_used", "mmm_months_used"),
    ("turbo_sessions_used", "turbo_sessions_used"),
    ("meta_downloads_used", "meta_downloads_used"),
)


async def _annual_portal_row_replace_from_excel(
    row: tuple,
    col_map: Dict[str, int],
    row_idx: int,
    cl: Dict[str, Any],
    errors: List[str],
) -> Optional[Tuple[Dict[str, Any], Dict[str, Any]]]:
    """
    Columns present in the upload header are the source of truth for that row.
    Empty cells clear subscription scalars, household key, and usage source;
    blank PRIMARY is treated as N; usage counters in mapped columns default to 0 when blank.
    Columns not present in the sheet leave existing DB values unchanged.
    """
    cid = str(cl.get("id") or "")
    prev_sub: Dict[str, Any] = dict(cl.get("annual_subscription") or {})
    new_sub: Dict[str, Any] = dict(prev_sub)
    new_sub.pop("awrp_year_label", None)
    client_extra: Dict[str, Any] = {}

    if "household_key" in col_map:
        hk_raw = _annual_upload_cell_str(row, col_map["household_key"])
        hk = (hk_raw or "").strip()[:200]
        client_extra["household_key"] = hk if hk else None

    if "is_primary_household_contact" in col_map:
        prim_raw = _annual_upload_cell_str(row, col_map["is_primary_household_contact"])
        pt = prim_raw.strip().upper()
        if not pt or pt in ("N", "NO", "0", "FALSE", "-", "—"):
            client_extra["is_primary_household_contact"] = False
        elif pt in ("Y", "YES", "1", "TRUE"):
            client_extra["is_primary_household_contact"] = True
        else:
            pl = prim_raw.strip().lower()
            if pl in ("y", "yes", "1", "true"):
                client_extra["is_primary_household_contact"] = True
            elif pl in ("n", "no", "0", "false"):
                client_extra["is_primary_household_contact"] = False
            else:
                errors.append(
                    f"Row {row_idx}: PRIMARY must be Y or N (got {prim_raw!r})"
                )
                return None

    if "annual_diid" in col_map:
        ad_raw = _annual_diid_cell_raw(row, col_map["annual_diid"]).strip()
        if not ad_raw:
            new_sub.pop("annual_diid", None)
        else:
            ad: Optional[str] = None
            ad_norm = normalize_annual_diid(ad_raw)
            if validate_annual_diid_format(ad_norm):
                ad = ad_norm
            elif re.fullmatch(r"\d{4}", ad_raw):
                nm = _display_name_for_portal_upload(row, col_map, cl)
                if not nm:
                    errors.append(
                        f"Row {row_idx}: for YYMM-only DIID {ad_raw!r}, set the Name column (or ensure the client has a name)"
                    )
                    return None
                built = _build_annual_diid_from_name_yymm(nm, ad_raw)
                if not built:
                    errors.append(
                        f"Row {row_idx}: invalid YYMM in {ad_raw!r} (month must be 01–12)"
                    )
                    return None
                ad = built
            else:
                errors.append(
                    f"Row {row_idx}: Annual DIID must be 8 characters (4 letters + YYMM) or 4 digits YYMM only (got {ad_raw!r})"
                )
                return None
            prev_ad = str(prev_sub.get("annual_diid") or "").strip()
            if ad != prev_ad:
                dup = await db.clients.find_one(
                    {
                        "annual_subscription.annual_diid": ad,
                        "id": {"$ne": cid},
                    }
                )
                if dup:
                    errors.append(
                        f"Row {row_idx}: DIID {ad} already used by another client"
                    )
                    return None
            new_sub["annual_diid"] = ad

    if "package_sku" in col_map:
        pkg_raw = _annual_upload_cell_str(row, col_map["package_sku"]).strip()
        if not pkg_raw:
            new_sub.pop("package_sku", None)
        else:
            sku = _parse_package_cell_for_upload(pkg_raw)
            if sku == "__invalid__":
                errors.append(
                    f"Row {row_idx}: package must be Home Coming (got {pkg_raw!r})"
                )
                return None
            new_sub["package_sku"] = sku

    if "start_date" in col_map:
        i_sd = col_map["start_date"]
        v_sd = row[i_sd] if i_sd < len(row) else None
        if v_sd in (None, ""):
            new_sub.pop("start_date", None)
        else:
            d_ok, err = _coerce_upload_date_value(v_sd, "start date")
            if err:
                errors.append(f"Row {row_idx}: {err}")
                return None
            new_sub["start_date"] = d_ok

    if "end_date" in col_map:
        i_ed = col_map["end_date"]
        v_ed = row[i_ed] if i_ed < len(row) else None
        if v_ed in (None, ""):
            new_sub.pop("end_date", None)
        else:
            d_ok, err = _coerce_upload_date_value(v_ed, "end date")
            if err:
                errors.append(f"Row {row_idx}: {err}")
                return None
            new_sub["end_date"] = d_ok

    if any(spec in col_map for _, spec in ANNUAL_PORTAL_USAGE_SPECS):
        prev_u = dict(new_sub.get("usage") or {})
        for u_key, spec in ANNUAL_PORTAL_USAGE_SPECS:
            if spec in col_map:
                cell = _annual_upload_cell_str(row, col_map[spec])
                if not cell:
                    n = 0
                else:
                    n, ierr = _parse_int_cell_strict(cell)
                    if ierr:
                        errors.append(f"Row {row_idx}: {spec} {ierr}")
                        return None
                prev_u[u_key] = n
        new_sub["usage"] = prev_u

    if "usage_source" in col_map:
        us = _annual_upload_cell_str(row, col_map["usage_source"]).strip()
        if not us:
            new_sub.pop("usage_source", None)
        else:
            t = us.lower()
            if t not in ("manual", "system"):
                errors.append(
                    f"Row {row_idx}: usage source must be manual or system"
                )
                return None
            new_sub["usage_source"] = t

    return new_sub, client_extra


# Admin Annual + dashboard: Excel header row — must stay aligned with upload column aliases in ANNUAL_PORTAL_UPLOAD_SPECS.
ANNUAL_PORTAL_EXCEL_HEADER_LABELS = [
    "#",
    "Name",
    "Email Id",
    "Start Date",
    "End Date",
    "DIID",
    "HomeComing",
    "AWRP months used",
    "MMM months used",
    "Turbo sessions used",
    "Meta downloads used",
    "Usage source",
    "HOUSEHOLD",
    "PRIMARY",
    "Client id",
]


@router.get("/annual-portal-subscribers/export")
async def download_annual_portal_subscription_export():
    """Current Annual + dashboard rows in the same shape as the template — edit and re-upload.

    Path uses two segments so it is never captured by ``GET /{client_id}`` (single-segment
    ``/annual-portal-subscription-export`` was matched as a client id → 404).
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    query = {"annual_member_dashboard": True}
    proj = {
        "_id": 0,
        "id": 1,
        "name": 1,
        "email": 1,
        "household_key": 1,
        "is_primary_household_contact": 1,
        "annual_subscription": 1,
        "portal_login_allowed": 1,
    }
    clients_list = await db.clients.find(query, proj).sort([("name", 1)]).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Annual portal"
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
    for col_idx, lab in enumerate(ANNUAL_PORTAL_EXCEL_HEADER_LABELS, 1):
        c = ws.cell(row=1, column=col_idx, value=lab)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    for row_idx, cl in enumerate(clients_list, start=2):
        sub = cl.get("annual_subscription") or {}
        usage = sub.get("usage") or {}
        pkg = (sub.get("package_sku") or "").strip().lower()
        home_cell = "Home Coming" if pkg == HOME_COMING_SKU else (sub.get("package_sku") or "")
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=(cl.get("name") or "").strip() or None)
        ws.cell(row=row_idx, column=3, value=(cl.get("email") or "").strip() or None)
        ws.cell(row=row_idx, column=4, value=(sub.get("start_date") or "").strip() or None)
        ws.cell(row=row_idx, column=5, value=(sub.get("end_date") or "").strip() or None)
        ws.cell(row=row_idx, column=6, value=(sub.get("annual_diid") or "").strip() or None)
        ws.cell(row=row_idx, column=7, value=home_cell or None)
        ws.cell(row=row_idx, column=8, value=int(usage.get("awrp_months_used") or 0))
        ws.cell(row=row_idx, column=9, value=int(usage.get("mmm_months_used") or 0))
        ws.cell(row=row_idx, column=10, value=int(usage.get("turbo_sessions_used") or 0))
        ws.cell(row=row_idx, column=11, value=int(usage.get("meta_downloads_used") or 0))
        us_src = (sub.get("usage_source") or "").strip().lower()
        ws.cell(row=row_idx, column=12, value=us_src if us_src in ("manual", "system") else None)
        ws.cell(row=row_idx, column=13, value=(cl.get("household_key") or "").strip() or None)
        ws.cell(
            row=row_idx,
            column=14,
            value="Y" if cl.get("is_primary_household_contact") else "N",
        )
        ws.cell(row=row_idx, column=15, value=cl.get("id") or None)

    widths = [5, 18, 26, 12, 12, 12, 14, 10, 10, 10, 10, 10, 22, 8, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    fname = f"annual_portal_subscribers_{ts}.xlsx"
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/annual-portal-subscription-template")
async def download_annual_portal_subscription_template():
    """Excel for bulk Home Coming annual_subscription updates — match by Client id and/or email."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    labels = ANNUAL_PORTAL_EXCEL_HEADER_LABELS
    wb = Workbook()
    ws = wb.active
    ws.title = "Annual portal"
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
    for col_idx, lab in enumerate(labels, 1):
        c = ws.cell(row=1, column=col_idx, value=lab)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    sample_primary = [
        "1",
        "Jane Primary",
        "primary@example.com",
        "2025-04-01",
        "2026-03-31",
        "JADO2504",
        "Home Coming",
        "3",
        "1",
        "0",
        "0",
        "manual",
        "Poonam Rathee",
        "Y",
        "",
    ]
    # Row 3: DIID as YYMM only (4 digits) — server builds 4 letters from Name + YYMM
    sample_peer = [
        "2",
        "Child Member",
        "",
        "2025-04-01",
        "2026-03-31",
        "2503",
        "Home Coming",
        "0",
        "0",
        "0",
        "0",
        "manual",
        "Poonam Rathee",
        "N",
        "paste-uuid-from-admin-grid",
    ]
    note_font = Font(italic=True, color="666666", size=10)
    for col_idx, val in enumerate(sample_primary, 1):
        c = ws.cell(row=2, column=col_idx, value=val)
        c.font = note_font
    for col_idx, val in enumerate(sample_peer, 1):
        c = ws.cell(row=3, column=col_idx, value=val)
        c.font = note_font
    widths = [5, 18, 26, 12, 12, 12, 14, 10, 10, 10, 10, 10, 22, 8, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=annual_portal_subscription_template.xlsx"
        },
    )


@router.post("/annual-portal-subscription-upload")
async def upload_annual_portal_subscription_excel(file: UploadFile = File(...)):
    """
    Replace annual_subscription / household fields from the sheet per row.
    Match by Client id, Email, or Name (+ optional HOUSEHOLD). Unmatched Name+HOUSEHOLD rows can
    create a new client (generated UUID). DIID may be full FFLlYYMM or YYMM only (letters from Name).
    Import sets annual_member_dashboard for updated clients. portal_login_allowed False still skips.
    """
    import io
    from openpyxl import load_workbook

    fn = (file.filename or "").lower()
    if not fn.endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Upload a .xlsx file (use Download template).",
        )
    raw = await file.read()
    if len(raw) > 8 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 8 MB).")

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read Excel file.")

    ws = wb.active
    header_row_idx, header_row = _find_annual_portal_header_row(ws)
    col_map = _annual_portal_upload_col_map(header_row)
    if (
        "email" not in col_map
        and "client_id" not in col_map
        and "row_name" not in col_map
    ):
        raise HTTPException(
            status_code=400,
            detail="Sheet must have an Email column, Client id column, and/or Name column (see template). If row 1 is a title, put headers on the next row.",
        )

    matched_labels = {
        field: str(header_row[idx] or "").strip() or field
        for field, idx in col_map.items()
        if idx < len(header_row)
    }

    updated = 0
    clients_created = 0
    skipped_blank = 0
    skipped_no_data = 0
    errors: List[str] = []
    max_rows = 5000
    row_count = 0
    data_start = header_row_idx + 1
    _template_sample_ids = frozenset({"paste-uuid-from-admin-grid"})
    _template_sample_emails = frozenset({"client@example.com", "primary@example.com"})

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start, values_only=True),
        start=data_start,
    ):
        row_count += 1
        if row_count > max_rows:
            errors.append(f"Stopped after {max_rows} data rows.")
            break
        if not row or not any(v not in (None, "") for v in row):
            skipped_blank += 1
            continue

        cid_cell = _annual_upload_cell_str(row, col_map.get("client_id")).strip()
        email = normalize_email(_annual_upload_cell_str(row, col_map.get("email")))
        if cid_cell in _template_sample_ids or (
            not cid_cell and email in _template_sample_emails
        ):
            skipped_blank += 1
            continue

        cl: Optional[Dict[str, Any]] = None
        if cid_cell:
            cl = await db.clients.find_one({"id": cid_cell})
            if not cl:
                errors.append(f"Row {row_idx}: no client with Client id {cid_cell}")
                continue
            if email and normalize_email(cl.get("email") or "") != email:
                errors.append(
                    f"Row {row_idx}: Email does not match this Client id — fix or clear Email"
                )
                continue
        elif email:
            cl = await db.clients.find_one({"email": email})
            if not cl:
                errors.append(f"Row {row_idx}: no client with email {email}")
                continue
        else:
            nm = _annual_upload_cell_str(row, col_map.get("row_name")).strip()
            if not nm:
                errors.append(
                    f"Row {row_idx}: set Client id, Email, or Name — all empty for this row"
                )
                continue
            hk_cell = ""
            if "household_key" in col_map:
                hk_cell = _annual_upload_cell_str(row, col_map["household_key"]).strip()
            if hk_cell:
                cl_nh, ehk = await _find_client_by_name_and_household(nm, hk_cell)
                if ehk:
                    errors.append(f"Row {row_idx}: {ehk}")
                    continue
                cl = cl_nh
                if cl is None:
                    created = await _create_client_for_annual_portal_upload(
                        row, col_map, row_idx, nm, errors
                    )
                    if created is None:
                        continue
                    cl = created
                    clients_created += 1
            else:
                by_n, nerr = await _find_single_client_by_name(nm)
                if nerr:
                    errors.append(f"Row {row_idx}: {nerr}")
                    continue
                cl = by_n
                if cl is None:
                    created = await _create_client_for_annual_portal_upload(
                        row, col_map, row_idx, nm, errors
                    )
                    if created is None:
                        continue
                    cl = created
                    clients_created += 1

        row_label = email or (cl.get("email") or "") or cid_cell or (cl.get("name") or "")
        if cl.get("portal_login_allowed") is False:
            errors.append(
                f"Row {row_idx}: {row_label} has portal blocked — skipped"
            )
            continue

        cid = cl["id"]
        if not (set(col_map.keys()) & ANNUAL_PORTAL_UPLOAD_PAYLOAD_KEYS):
            skipped_no_data += 1
            continue

        built = await _annual_portal_row_replace_from_excel(
            row, col_map, row_idx, cl, errors
        )
        if built is None:
            continue
        new_sub, client_extra = built

        now = datetime.now(timezone.utc).isoformat()
        set_doc: Dict[str, Any] = {
            "annual_subscription": new_sub,
            "annual_member_dashboard": True,
            "updated_at": now,
        }
        set_doc.update(client_extra)
        await db.clients.update_one({"id": cid}, {"$set": set_doc})
        updated += 1

    return {
        "updated": updated,
        "clients_created": clients_created,
        "skipped_blank_rows": skipped_blank,
        "skipped_no_data_rows": skipped_no_data,
        "skipped_empty_rows": skipped_blank + skipped_no_data,
        "header_row": header_row_idx,
        "matched_columns": matched_labels,
        "errors": errors[:100],
        "error_count": len(errors),
    }


class BulkSetPortalLoginBody(BaseModel):
    client_ids: List[str]
    portal_login_allowed: bool


@router.post("/bulk-set-portal-login")
async def bulk_set_portal_login(data: BulkSetPortalLoginBody):
    """
    Set portal_login_allowed for many clients in one request.
    When enabling (True), sends the welcome email for each client that was explicitly blocked (False → True).
    """
    raw_ids = [str(i).strip() for i in (data.client_ids or []) if i is not None and str(i).strip()]
    # Dedupe while preserving order
    seen = set()
    ids = []
    for i in raw_ids:
        if i not in seen:
            seen.add(i)
            ids.append(i)
    if not ids:
        raise HTTPException(status_code=400, detail="No client_ids provided")
    if len(ids) > 500:
        raise HTTPException(status_code=400, detail="Maximum 500 clients per request")

    value = bool(data.portal_login_allowed)
    now = datetime.now(timezone.utc).isoformat()
    updated = 0
    not_found = 0
    welcome_emails_sent = 0
    welcome_emails_failed = 0

    for cid in ids:
        cl = await db.clients.find_one({"id": cid})
        if not cl:
            not_found += 1
            continue

        set_doc: Dict[str, Any] = {"portal_login_allowed": value, "updated_at": now}
        if value:
            set_doc["intake_pending"] = False
        await db.clients.update_one({"id": cid}, {"$set": set_doc})
        updated += 1

        if value and cl.get("portal_login_allowed") is False:
            to_em = (cl.get("email") or "").strip()
            if to_em:
                try:
                    from routes.emails import send_dashboard_access_granted_email

                    ok = await send_dashboard_access_granted_email(to_em, cl.get("name") or "")
                    if ok:
                        welcome_emails_sent += 1
                    else:
                        welcome_emails_failed += 1
                except Exception:
                    welcome_emails_failed += 1
                    logger.exception("bulk_set_portal_login: email failed for client_id=%s", cid)

    return {
        "updated": updated,
        "not_found": not_found,
        "welcome_emails_sent": welcome_emails_sent,
        "welcome_emails_failed": welcome_emails_failed,
    }


class BulkDashboardAccessBody(BaseModel):
    """Only fields present in the JSON body are applied to every selected client."""

    client_ids: List[str]
    annual_member_dashboard: Optional[bool] = None
    preferred_payment_method: Optional[str] = None
    india_payment_method: Optional[str] = None
    india_discount_percent: Optional[float] = None
    india_tax_enabled: Optional[bool] = None
    india_tax_percent: Optional[float] = None
    india_tax_label: Optional[str] = None
    preferred_india_gpay_id: Optional[str] = None
    preferred_india_bank_id: Optional[str] = None


@router.post("/bulk-update-dashboard-access")
async def bulk_update_dashboard_access(data: BulkDashboardAccessBody):
    """Apply the same dashboard-access fields (access type, payments, GST, discount) to many clients."""
    raw_ids = [str(i).strip() for i in (data.client_ids or []) if i is not None and str(i).strip()]
    seen = set()
    ids = []
    for i in raw_ids:
        if i not in seen:
            seen.add(i)
            ids.append(i)
    if not ids:
        raise HTTPException(status_code=400, detail="No client_ids provided")
    if len(ids) > 500:
        raise HTTPException(status_code=400, detail="Maximum 500 clients per request")

    patch = data.model_dump(exclude_unset=True)
    patch.pop("client_ids", None)
    if not patch:
        raise HTTPException(status_code=400, detail="No fields to update — send at least one field besides client_ids")

    uf: Dict[str, Any] = {}
    if "annual_member_dashboard" in patch:
        uf["annual_member_dashboard"] = bool(patch["annual_member_dashboard"])
    if "preferred_payment_method" in patch:
        pm = (patch.get("preferred_payment_method") or "").strip().lower()
        uf["preferred_payment_method"] = pm if pm else None
    if "india_payment_method" in patch:
        vpm = (patch.get("india_payment_method") or "").strip()
        uf["india_payment_method"] = vpm if vpm else None
    if "india_discount_percent" in patch:
        idp = patch.get("india_discount_percent")
        if idp is None:
            uf["india_discount_percent"] = None
        else:
            uf["india_discount_percent"] = float(idp)
    if "india_tax_enabled" in patch:
        uf["india_tax_enabled"] = bool(patch["india_tax_enabled"])
    if "india_tax_percent" in patch:
        itp = patch.get("india_tax_percent")
        uf["india_tax_percent"] = float(itp) if itp is not None else None
    if "india_tax_label" in patch:
        uf["india_tax_label"] = (patch.get("india_tax_label") or "GST").strip() or "GST"
    if "preferred_india_gpay_id" in patch:
        uf["preferred_india_gpay_id"] = (patch.get("preferred_india_gpay_id") or "").strip()
    if "preferred_india_bank_id" in patch:
        uf["preferred_india_bank_id"] = (patch.get("preferred_india_bank_id") or "").strip()

    now = datetime.now(timezone.utc).isoformat()
    uf["updated_at"] = now
    updated = 0
    not_found = 0

    for cid in ids:
        cl = await db.clients.find_one({"id": cid})
        if not cl:
            not_found += 1
            continue
        await db.clients.update_one({"id": cid}, {"$set": uf})
        updated += 1

    return {"updated": updated, "not_found": not_found}


@router.get("/{client_id}")
async def get_client(client_id: str):
    cl = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not cl:
        raise HTTPException(status_code=404, detail="Client not found")
    cl["first_program"] = effective_first_program(cl)
    return cl


# ========== UPDATE ==========


class IndiaDiscountMemberBand(BaseModel):
    """India checkout discount for participant count in [min, max] (inclusive).

    Use either ``percent`` or ``amount`` (fixed INR off the effective base), not both.
    Legacy stored key ``amount_inr`` is accepted on input and normalized to ``amount``.
    """

    min: int = Field(ge=0, le=999)
    max: int = Field(ge=0, le=999)
    percent: Optional[float] = Field(default=None, ge=0, le=100)
    amount: Optional[float] = Field(default=None, ge=0, le=1_000_000_000)

    @model_validator(mode="before")
    @classmethod
    def normalize_legacy_amount_inr(cls, data: Any) -> Any:
        if isinstance(data, dict):
            d = dict(data)
            if d.get("amount_inr") is not None and d.get("amount") is None:
                d["amount"] = d.pop("amount_inr")
            return d
        return data

    @model_validator(mode="after")
    def validate_range_and_discount(self) -> "IndiaDiscountMemberBand":
        if self.max < self.min:
            raise ValueError("max must be >= min")
        has_amt = self.amount is not None and float(self.amount) > 0
        has_pct = self.percent is not None and float(self.percent) > 0
        if has_amt and has_pct:
            raise ValueError("Use either amount or percent, not both")
        if not has_amt and not has_pct:
            raise ValueError("Set a positive percent or amount on each band")
        return self


class ClientUpdate(BaseModel):
    label_manual: Optional[str] = None
    diid_middle: Optional[str] = None  # 4 letters + YYMM between DIID- and hex suffix
    first_program_manual: Optional[str] = None  # optional CRM override for first program column
    notes: Optional[str] = None
    name: Optional[str] = None
    phone: Optional[str] = None
    immediate_family_editing_approved: Optional[bool] = None
    # When False, Google / student portal sign-in is blocked until set to True.
    portal_login_allowed: Optional[bool] = None
    india_tax_enabled: Optional[bool] = None
    india_tax_percent: Optional[float] = None
    india_tax_label: Optional[str] = None
    india_tax_visible_on_dashboard: Optional[bool] = None
    india_payment_method: Optional[str] = None   # e.g. "gpay", "upi", "bank_transfer", "any"
    india_discount_percent: Optional[float] = None  # client-specific discount on base price
    # When set, first matching band (by order) overrides india_discount_percent for that participant count.
    india_discount_member_bands: Optional[List[IndiaDiscountMemberBand]] = None
    preferred_payment_method: Optional[str] = None  # gpay_upi, bank_transfer, cash_deposit, stripe (intake / CRM)
    intake_pending: Optional[bool] = None
    family_pending_review: Optional[bool] = None
    family_approved: Optional[bool] = None        # once True, list is permanently frozen
    # India proof row tags (merged with subscription.* in student home when subscription is empty)
    preferred_india_gpay_id: Optional[str] = None
    preferred_india_bank_id: Optional[str] = None
    # When True, Sacred Home uses annual-subscriber pricing like an Excel-tagged Iris member
    annual_member_dashboard: Optional[bool] = None
    # CRM: same key on each family member’s client row; optional primary flag on the manager’s row
    household_key: Optional[str] = None
    is_primary_household_contact: Optional[bool] = None
    email: Optional[str] = None
    discovery_source: Optional[str] = None
    discovery_other_note: Optional[str] = None
    referred_by_client_id: Optional[str] = None
    # CRM grid first-seen column: YYYY-MM (1st of month), YYYY-MM-DD, or ISO 8601 → stored as UTC ISO on ``created_at``
    created_at: Optional[str] = None
    # Portal cohort for layered AWRP / batch pricing (Admin → Dashboard settings)
    awrp_batch_id: Optional[str] = None


@router.put("/{client_id}")
async def update_client(client_id: str, data: ClientUpdate):
    cl = await db.clients.find_one({"id": client_id})
    if not cl:
        raise HTTPException(status_code=404, detail="Client not found")

    update_fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
    incoming = data.model_dump(exclude_unset=True)
    if data.label_manual is not None:
        lm = (data.label_manual or "").strip()
        if lm:
            if not is_allowed_manual_label(lm):
                raise HTTPException(status_code=400, detail="Invalid garden label")
            lm = normalize_label(lm)
        update_fields["label_manual"] = lm
        if lm:
            update_fields["label"] = lm
    if data.notes is not None:
        update_fields["notes"] = data.notes
    if data.name is not None:
        nm = str(data.name).strip()
        update_fields["name"] = normalize_person_name(nm) if nm else ""
    if "created_at" in incoming and data.created_at is not None:
        update_fields["created_at"] = _coerce_client_created_at_iso(str(data.created_at))
    if "phone" in incoming:
        pn = normalize_phone(str(data.phone) if data.phone is not None else "")
        update_fields["phone"] = pn or None
    if "diid_middle" in incoming and str(data.diid_middle or "").strip():
        update_fields["diid"] = await _new_diid_with_middle(
            db, client_id, cl, str(data.diid_middle).strip()
        )
    if "first_program_manual" in incoming:
        fp = (data.first_program_manual or "").strip() if data.first_program_manual is not None else ""
        update_fields["first_program_manual"] = fp if fp else None

    if "discovery_source" in incoming:
        ds_in = data.discovery_source
        if ds_in is None or (isinstance(ds_in, str) and not str(ds_in).strip()):
            update_fields["discovery_source"] = None
        else:
            update_fields["discovery_source"] = normalize_discovery_source_value(str(ds_in))
    if "discovery_other_note" in incoming:
        on = (
            (data.discovery_other_note or "").strip()
            if data.discovery_other_note is not None
            else ""
        )
        update_fields["discovery_other_note"] = (on[:500] if on else None)

    eff_disc = update_fields.get("discovery_source", cl.get("discovery_source"))
    if "discovery_source" in incoming and eff_disc != "Referral":
        update_fields["referred_by_client_id"] = None
        update_fields["referred_by_name"] = None

    if "referred_by_client_id" in incoming:
        raw_rb = data.referred_by_client_id
        eff_disc = update_fields.get("discovery_source", cl.get("discovery_source"))
        if raw_rb is None or (isinstance(raw_rb, str) and not str(raw_rb).strip()):
            update_fields["referred_by_client_id"] = None
            update_fields["referred_by_name"] = None
        elif eff_disc != "Referral":
            raise HTTPException(
                status_code=400,
                detail="Set How they found us to Referral before saving a referrer UUID.",
            )
        else:
            rid, rname = await resolve_referrer_client(db, client_id, str(raw_rb))
            update_fields["referred_by_client_id"] = rid
            update_fields["referred_by_name"] = rname

    if data.immediate_family_editing_approved is not None:
        update_fields["immediate_family_editing_approved"] = bool(data.immediate_family_editing_approved)
    if data.portal_login_allowed is not None:
        update_fields["portal_login_allowed"] = bool(data.portal_login_allowed)
    if data.india_tax_enabled is not None:
        update_fields["india_tax_enabled"] = bool(data.india_tax_enabled)
    if data.india_tax_percent is not None:
        update_fields["india_tax_percent"] = float(data.india_tax_percent)
    if data.india_tax_label is not None:
        update_fields["india_tax_label"] = data.india_tax_label
    if data.india_tax_visible_on_dashboard is not None:
        update_fields["india_tax_visible_on_dashboard"] = bool(data.india_tax_visible_on_dashboard)
    if data.india_payment_method is not None:
        vpm = (data.india_payment_method or "").strip()
        update_fields["india_payment_method"] = vpm if vpm else None
    if data.india_discount_percent is not None:
        update_fields["india_discount_percent"] = float(data.india_discount_percent)
    if data.preferred_payment_method is not None:
        pm = (data.preferred_payment_method or "").strip().lower()
        update_fields["preferred_payment_method"] = pm if pm else None
    if data.intake_pending is not None:
        update_fields["intake_pending"] = bool(data.intake_pending)
    if data.family_pending_review is not None:
        update_fields["family_pending_review"] = bool(data.family_pending_review)
    if data.family_approved is not None:
        update_fields["family_approved"] = bool(data.family_approved)
        if data.family_approved:
            # Approve & Freeze: lock the list, clear pending flag, revoke any re-edit approval
            update_fields["immediate_family_locked"] = True
            update_fields["family_pending_review"] = False
            update_fields["immediate_family_editing_approved"] = False
    if data.preferred_india_gpay_id is not None:
        update_fields["preferred_india_gpay_id"] = (data.preferred_india_gpay_id or "").strip()
    if data.preferred_india_bank_id is not None:
        update_fields["preferred_india_bank_id"] = (data.preferred_india_bank_id or "").strip()
    if data.annual_member_dashboard is not None:
        update_fields["annual_member_dashboard"] = bool(data.annual_member_dashboard)
    if data.household_key is not None:
        hk = (data.household_key or "").strip()[:200]
        update_fields["household_key"] = hk if hk else None
    if data.is_primary_household_contact is not None:
        update_fields["is_primary_household_contact"] = bool(data.is_primary_household_contact)
    if "awrp_batch_id" in incoming:
        ab = data.awrp_batch_id
        if ab is None or (isinstance(ab, str) and not str(ab).strip()):
            update_fields["awrp_batch_id"] = None
        else:
            update_fields["awrp_batch_id"] = str(ab).strip()

    if data.email is not None:
        new_em = normalize_email(data.email)
        if new_em and "@" not in new_em:
            raise HTTPException(status_code=400, detail="Invalid email address")
        old_em = normalize_email(cl.get("email") or "")
        if new_em != old_em:
            update_fields["email"] = new_em or None

    if "india_discount_member_bands" in incoming:
        bands = data.india_discount_member_bands
        if bands:
            update_fields["india_discount_member_bands"] = [
                b.model_dump(exclude_none=True) for b in bands
            ]
        else:
            update_fields["india_discount_member_bands"] = None

    # New-intake queue clears when Google login is enabled (no separate "mark reviewed" needed)
    if data.portal_login_allowed is not None and bool(data.portal_login_allowed):
        update_fields["intake_pending"] = False

    await db.clients.update_one({"id": client_id}, {"$set": update_fields})

    if "email" in update_fields:
        synced = (update_fields["email"] or "").strip()
        if synced:
            await db.users.update_many(
                {"client_id": client_id},
                {"$set": {"email": synced, "updated_at": update_fields["updated_at"]}},
            )

    if "name" in update_fields:
        await db.users.update_many(
            {"client_id": client_id},
            {
                "$set": {
                    "name": update_fields.get("name") or "",
                    "updated_at": update_fields["updated_at"],
                }
            },
        )

    # Email when Google / student portal access is newly enabled (was blocked, e.g. after intake)
    if (
        data.portal_login_allowed is not None
        and bool(data.portal_login_allowed) is True
        and cl.get("portal_login_allowed") is False
    ):
        merged = update_fields.get("email") if "email" in update_fields else cl.get("email")
        to_em = (merged or "").strip() if merged is not None else ""
        if to_em:
            try:
                from routes.emails import send_dashboard_access_granted_email

                greet_name = (
                    update_fields.get("name") if "name" in update_fields else cl.get("name")
                ) or ""
                await send_dashboard_access_granted_email(to_em, greet_name)
            except Exception:
                logger.exception("send_dashboard_access_granted_email failed for client_id=%s", client_id)

    # Back to automatic rules (usually Dew until paid conversions) when override cleared
    if data.label_manual is not None and not (data.label_manual or "").strip():
        updated = await db.clients.find_one({"id": client_id}, {"_id": 0})
        new_label = await compute_label(updated)
        await db.clients.update_one({"id": client_id}, {"$set": {"label": new_label}})

    return {"message": "Updated"}


# ========== DELETE ==========

@router.delete("/{client_id}")
async def delete_client(client_id: str):
    result = await db.clients.delete_one({"id": client_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    return {"message": "Deleted"}


# ========== DOWNLOAD ==========

@router.get("/export/csv")
async def export_clients_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    clients_list = await db.clients.find({}, {"_id": 0}).sort("label", 1).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Garden"

    headers = [
        "DIID",
        "UUID (internal id)",
        "Legacy DID",
        "Label",
        "First program joined",
        "Name",
        "How found us",
        "Other (detail)",
        "Referrer UUID",
        "Referrer name",
        "Email",
        "Phone",
        "Household key",
        "Primary household contact",
        "Annual program",
        "Sources",
        "Programs Enrolled",
        "Total Conversions",
        "First Contact",
        "Last Updated",
        "Notes",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="E8E0C8"))

    label_fills = {
        "dew": PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"),
        "seed": PatternFill(start_color="ECFCCB", end_color="ECFCCB", fill_type="solid"),
        "root": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "bloom": PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid"),
        "iris": PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"),
        "purple_bees": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
        "iris_bees": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
    }

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, cl in enumerate(clients_list, 2):
        programs = ", ".join(set(c.get("program_title", "") for c in cl.get("conversions", []) if c.get("program_title")))
        sources = ", ".join(set(cl.get("sources", [])))
        label = cl.get("label") or LABEL_DEW
        first_prog = effective_first_program(cl)
        row_data = [
            cl.get("diid", ""),
            cl.get("id", ""),
            cl.get("did", ""),
            label,
            first_prog,
            cl.get("name", ""),
            cl.get("discovery_source") or "",
            cl.get("discovery_other_note") or "",
            cl.get("referred_by_client_id") or "",
            cl.get("referred_by_name") or "",
            cl.get("email", ""),
            cl.get("phone", ""),
            cl.get("household_key") or "",
            "Yes" if cl.get("is_primary_household_contact") else "",
            "Yes" if cl.get("annual_member_dashboard") else "No",
            sources,
            programs,
            len(cl.get("conversions", [])),
            cl.get("created_at", ""),
            cl.get("updated_at", ""),
            cl.get("notes", ""),
        ]

        fill = label_fills.get(label_stripe_key(label), PatternFill())
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col, value=str(val) if val else "")
            cell.fill = fill
            cell.border = thin_border

    col_widths = [30, 38, 14, 52, 28, 20, 14, 24, 36, 22, 30, 18, 22, 12, 14, 25, 40, 16, 22, 22, 30]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[ws.cell(row=1, column=i + 1).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=divine_iris_clients_{timestamp}.xlsx"}
    )
