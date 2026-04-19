"""
Annual Subscribers — completely isolated from the clients collection.
Reads/writes only the `annual_subscribers` collection.
Zero impact on live dashboard, client records, discounts, or payment flows.
"""
from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timezone
import os
import uuid
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent.parent
load_dotenv(ROOT_DIR / '.env')

router = APIRouter(prefix="/api/annual-subscribers", tags=["Annual Subscribers"])

mongo_url = os.environ['MONGO_URL']
_client = AsyncIOMotorClient(mongo_url)
db = _client[os.environ['DB_NAME']]
# Dedicated collection — never touches db.clients
col = db["annual_subscribers"]

COLUMNS = [
    ("email",                        "Email *"),
    ("name",                         "Name"),
    ("phone",                        "Phone"),
    ("did",                          "Divine Iris ID"),
    ("annual_start_date",            "Annual Start Date (YYYY-MM-DD)"),
    ("annual_end_date",              "Annual End Date (YYYY-MM-DD)"),
    ("portal_login_allowed",         "Portal Access (Yes/No, default Yes)"),
    ("india_payment_method",         "Payment Method (gpay/upi/bank_transfer/any)"),
    ("india_discount_percent",       "Discount % on Base"),
    ("india_tax_enabled",            "Tax Enabled (Yes/No)"),
    ("india_tax_percent",            "Tax %"),
    ("india_tax_label",              "Tax Label (e.g. GST)"),
    ("sponsorship_discount_percent", "Sponsorship Discount %"),
    ("notes",                        "Notes / Pause Reason"),
]


class AnnualSubscriberUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    did: Optional[str] = None
    annual_start_date: Optional[str] = None
    annual_end_date: Optional[str] = None
    portal_login_allowed: Optional[bool] = None
    india_payment_method: Optional[str] = None
    india_discount_percent: Optional[float] = None
    india_tax_enabled: Optional[bool] = None
    india_tax_percent: Optional[float] = None
    india_tax_label: Optional[str] = None
    sponsorship_discount_percent: Optional[float] = None
    notes: Optional[str] = None


# ── List ──────────────────────────────────────────────────────────────────────

@router.get("")
async def list_annual_subscribers():
    docs = await col.find({}, {"_id": 0}).sort("name", 1).to_list(5000)
    return docs


# ── Create ────────────────────────────────────────────────────────────────────

class AnnualSubscriberCreate(AnnualSubscriberUpdate):
    email: str


@router.post("")
async def create_annual_subscriber(data: AnnualSubscriberCreate):
    email = (data.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="email is required")
    existing = await col.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=409, detail="A record with this email already exists")
    doc = {
        "id": str(uuid.uuid4()),
        "email": email,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "portal_login_allowed": True,
        **{k: v for k, v in data.dict(exclude={"email"}).items() if v is not None},
    }
    await col.insert_one(doc)
    doc.pop("_id", None)
    return doc


# ── Update ────────────────────────────────────────────────────────────────────

@router.put("/{record_id}")
async def update_annual_subscriber(record_id: str, data: AnnualSubscriberUpdate):
    existing = await col.find_one({"id": record_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Record not found")
    fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
    for k, v in data.dict(exclude_none=True).items():
        fields[k] = v
    await col.update_one({"id": record_id}, {"$set": fields})
    return {"message": "Updated"}


# ── Delete ────────────────────────────────────────────────────────────────────

@router.delete("/{record_id}")
async def delete_annual_subscriber(record_id: str):
    result = await col.delete_one({"id": record_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Record not found")
    return {"message": "Deleted"}


# ── Excel template ────────────────────────────────────────────────────────────

@router.get("/excel-template")
async def download_template():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Annual Subscribers"

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    note_font = Font(italic=True, color="888888", size=9)
    note_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")

    for col_idx, (_, label) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    sample = [
        "jane@example.com", "Jane Doe", "+91 98765 43210", "DI-001",
        "2024-01-01", "2024-12-31", "Yes", "gpay",
        "15", "Yes", "18", "GST", "", "Active member",
    ]
    for col_idx, val in enumerate(sample, 1):
        c = ws.cell(row=2, column=col_idx, value=val)
        c.font = note_font
        c.fill = note_fill

    col_widths = [28, 22, 18, 14, 26, 24, 28, 28, 18, 20, 10, 16, 22, 30]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[ws.cell(row=1, column=i + 1).column_letter].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=annual_subscribers_template.xlsx"},
    )


# ── Bulk Excel upload ─────────────────────────────────────────────────────────

@router.post("/upload")
async def upload_excel(file: UploadFile = File(...)):
    """Upsert rows by email into the annual_subscribers collection only.
    Never touches db.clients.
    """
    import io
    from openpyxl import load_workbook

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    header_row = [str(c.value or "").strip() for c in ws[1]]
    col_map: dict[str, int] = {}
    for field, label in COLUMNS:
        clean_label = label.lower().replace("*", "").strip()
        for i, h in enumerate(header_row):
            if h.lower().replace("*", "").strip() in (clean_label, field.lower()):
                col_map[field] = i
                break

    created = updated = skipped = 0
    errors: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        def get(field: str) -> str:
            i = col_map.get(field)
            v = row[i] if i is not None and i < len(row) else None
            return str(v).strip() if v is not None else ""

        email = (get("email") or "").lower().strip()
        if not email:
            errors.append(f"Row {row_idx}: missing email, skipped")
            skipped += 1
            continue

        def parse_bool(val: str, default: bool = True) -> bool:
            if val == "":
                return default
            return val.lower() in ("yes", "true", "1", "y")

        def parse_float(val: str):
            try:
                return float(val) if val != "" else None
            except ValueError:
                return None

        fields: dict = {"updated_at": datetime.now(timezone.utc).isoformat()}

        portal_raw = get("portal_login_allowed")
        fields["portal_login_allowed"] = parse_bool(portal_raw, True)

        for f in ("name", "phone", "did", "annual_start_date", "annual_end_date",
                  "india_payment_method", "india_tax_label", "notes"):
            v = get(f)
            if v:
                fields[f] = v

        for f in ("india_discount_percent", "india_tax_percent", "sponsorship_discount_percent"):
            v = parse_float(get(f))
            if v is not None:
                fields[f] = v

        it = get("india_tax_enabled")
        if it:
            fields["india_tax_enabled"] = parse_bool(it, False)

        existing = await col.find_one({"email": email})
        if existing:
            await col.update_one({"email": email}, {"$set": fields})
            updated += 1
        else:
            new_doc = {
                "id": str(uuid.uuid4()),
                "email": email,
                "created_at": datetime.now(timezone.utc).isoformat(),
                **fields,
            }
            await col.insert_one(new_doc)
            created += 1

    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}
