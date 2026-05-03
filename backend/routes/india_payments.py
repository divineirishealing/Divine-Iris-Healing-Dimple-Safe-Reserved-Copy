import asyncio

from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from typing import Any, Dict, List, Optional, Tuple
import os, uuid, logging
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path
from datetime import datetime, timezone
import re
import mimetypes

import s3_storage
from routes.clients import ensure_client_from_enrollment_lead

ROOT_DIR = Path(__file__).parent.parent
load_dotenv(ROOT_DIR / '.env')

router = APIRouter(prefix="/api/india-payments", tags=["India Payments"])

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

logger = logging.getLogger(__name__)

# ─── Admin enrollment reporting: one `enrollments` collection for all flows ───


def _txn_sort_key(t: dict) -> float:
    v = t.get("updated_at") or t.get("created_at") or t.get("paid_at")
    if v is None:
        return 0.0
    if hasattr(v, "timestamp"):
        try:
            return float(v.timestamp())
        except Exception:
            return 0.0
    if isinstance(v, str):
        try:
            from datetime import datetime as dt

            return dt.fromisoformat(v.replace("Z", "+00:00")).timestamp()
        except Exception:
            return 0.0
    return 0.0


def _norm_cmp_id(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _pick_best_transaction_for_enrollment(enrollment: dict, txns: List[dict]) -> Optional[dict]:
    """Prefer paid txns that match this enrollment's catalog line, then highest amount (not merely latest touch)."""
    if not txns:
        return None

    def _is_paid(t: dict) -> bool:
        s = str(t.get("payment_status", "")).lower()
        return s in ("paid", "complete", "completed")

    paid_pool = [t for t in txns if _is_paid(t)]
    pool = paid_pool if paid_pool else list(txns)
    e_it = _norm_cmp_id(enrollment.get("item_id"))
    e_ty = _clean_str(enrollment.get("item_type")).lower()

    candidates = list(pool)
    if e_it:
        matched = [t for t in pool if _norm_cmp_id(t.get("item_id")) == e_it]
        if matched:
            candidates = matched
    if e_ty:
        ty_m = [t for t in candidates if _clean_str(t.get("item_type")).lower() == e_ty]
        if ty_m:
            candidates = ty_m

    def amt_key(t: dict) -> float:
        try:
            return float(t.get("amount") or 0)
        except (TypeError, ValueError):
            return 0.0

    return max(candidates, key=lambda t: (amt_key(t), _txn_sort_key(t)))


async def _transactions_grouped_by_enrollment(enrollment_ids: List[str]) -> Dict[str, List[dict]]:
    if not enrollment_ids:
        return {}
    all_tx = await db.payment_transactions.find(
        {"enrollment_id": {"$in": enrollment_ids}},
        {"_id": 0},
    ).to_list(20000)
    out: Dict[str, List[dict]] = {}
    for t in all_tx:
        eid = t.get("enrollment_id")
        if not eid:
            continue
        out.setdefault(eid, []).append(t)
    return out


def _enrollment_origin(e: dict) -> str:
    """Rough source: student dashboard annual flow vs public site (cart / program / session pages)."""
    if e.get("dashboard_checkout_ready") or e.get("dashboard_mixed_total") is not None:
        return "dashboard"
    return "website"


def _clean_str(val: Any) -> str:
    if val is None:
        return ""
    s = str(val)
    if s in ("None", "null"):
        return ""
    return s


def _scalar_field(d: dict, *keys: str) -> Any:
    """First non-empty value for any of the keys (exact or case-insensitive)."""
    if not isinstance(d, dict):
        return None
    lower_map = {str(k).lower(): v for k, v in d.items()}
    for k in keys:
        if k in d:
            v = d[k]
            if v is not None and v != "":
                return v
        lk = k.lower()
        if lk in lower_map:
            v = lower_map[lk]
            if v is not None and v != "":
                return v
    return None


def display_program_title_for_enrollment(
    enrollment: dict,
    participants: Optional[List[dict]] = None,
    catalog_item: Optional[dict] = None,
) -> str:
    """Label for admin/reporting: Home Coming cart title on participants wins over catalog program name."""
    et = _clean_str(enrollment.get("item_title"))
    cat = _clean_str((catalog_item or {}).get("title"))
    pool = participants if participants is not None else enrollment.get("participants")
    titles: List[str] = []
    if isinstance(pool, list):
        for p in pool:
            if not isinstance(p, dict):
                continue
            t = _clean_str(_scalar_field(p, "program_title", "programTitle"))
            if t:
                titles.append(t)
    for t in titles:
        if "home coming" in t.lower():
            return t
    if et:
        return et
    if titles:
        return titles[0]
    return cat


def _format_age_for_report(val: Any) -> str:
    if val is None or val == "":
        return ""
    try:
        if isinstance(val, float):
            return str(int(val)) if val == int(val) else str(val).strip()
        return str(int(val))
    except (TypeError, ValueError):
        s = str(val).strip()
        return s


def _merge_participants_for_report(enrollment: dict, txn: Optional[dict]) -> List[dict]:
    """
    Prefer enrollment.participants; merge with transaction.participants index-wise so
    missing fields on one side are filled from the other. If enrollment has no dict rows
    but the txn snapshot does, use txn (fixes sparse / legacy enrollments).
    """
    raw = [x for x in (enrollment.get("participants") or []) if isinstance(x, dict)]
    tx_list = [x for x in ((txn or {}).get("participants") or []) if isinstance(x, dict)]
    if not raw:
        return list(tx_list)
    if not tx_list:
        return list(raw)
    merged: List[dict] = []
    n = max(len(raw), len(tx_list))
    for i in range(n):
        a = raw[i] if i < len(raw) else {}
        b = tx_list[i] if i < len(tx_list) else {}
        m = {**b, **a}
        for k, v in b.items():
            if (m.get(k) in (None, "")) and v not in (None, ""):
                m[k] = v
        merged.append(m)
    return merged


def _notify_yes_from_participant(p: dict) -> bool:
    v = _scalar_field(p, "notify", "notify_enrollment", "Notify")
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "y", "on")


def _attendance_mode_from_participant(p: dict) -> str:
    v = _scalar_field(p, "attendance_mode", "attendanceMode", "session_mode", "mode")
    return _clean_str(v).lower() if v is not None and v != "" else ""


def _payment_amount_currency(e: dict, txn: Optional[dict]) -> tuple:
    """Display amount + currency; fall back to dashboard quote when no txn yet."""
    if txn:
        return txn.get("amount", 0) or 0, _clean_str(txn.get("currency")).lower()
    amt = e.get("dashboard_mixed_total")
    cur = _clean_str(e.get("dashboard_mixed_currency")).lower()
    if amt is not None:
        try:
            return float(amt), cur
        except Exception:
            return 0, cur
    return 0, cur


def _client_portal_cohort_from_client_doc(client: Optional[dict]) -> str:
    """Same precedence as subscriber/client editing: top-level awrp_batch_id, then subscription, then annual."""
    if not isinstance(client, dict):
        return ""
    for key in ("awrp_batch_id",):
        raw = client.get(key)
        if raw is not None and str(raw).strip():
            return str(raw).strip()
    sub = client.get("subscription") or {}
    raw = sub.get("awrp_batch_id")
    if raw is not None and str(raw).strip():
        return str(raw).strip()
    annual_sub = client.get("annual_subscription") or {}
    raw = annual_sub.get("awrp_batch_id")
    if raw is not None and str(raw).strip():
        return str(raw).strip()
    return ""


def _all_portal_client_ids_from_txns(txns_by_eid: Dict[str, List[dict]]) -> List[str]:
    out: List[str] = []
    seen = set()
    for txns in (txns_by_eid or {}).values():
        for t in txns or []:
            cid = _norm_cmp_id(t.get("portal_client_id"))
            if cid and cid not in seen:
                seen.add(cid)
                out.append(cid)
    return out


def _client_ids_for_emi_context(enrollments: List[dict], txns_by_eid: Dict[str, List[dict]]) -> List[str]:
    """Portal client ids from txns plus ``enrollment.client_id`` (dashboard Sacred Home often has the latter only)."""
    out: List[str] = []
    seen = set()

    def add(raw: Any) -> None:
        cid = _norm_cmp_id(raw)
        if cid and cid not in seen:
            seen.add(cid)
            out.append(cid)

    for e in enrollments or []:
        add((e or {}).get("client_id"))
        eid = (e or {}).get("id")
        for t in txns_by_eid.get(eid) or []:
            add((t or {}).get("portal_client_id"))
    return out


async def _booker_email_client_ids_for_emi_overlay(
    enrollments: List[dict],
    txns_by_eid: Dict[str, List[dict]],
) -> Tuple[List[str], Dict[str, str]]:
    """
    When ``enrollment.client_id`` is missing, map contact emails → ``clients.id`` so EMI tier tagging
    can still load ``subscription.emis``. Uses booker email on the enrollment plus payer / booker
    emails on linked payment rows (Sacred Home often stores the portal payer only on the txn).
    """
    emails: List[str] = []
    seen_em = set()

    def add_em(raw: Any) -> None:
        em = (_clean_str(raw) or "").strip().lower()
        if em and em not in seen_em:
            seen_em.add(em)
            emails.append(em)

    for e in enrollments or []:
        if _norm_cmp_id((e or {}).get("client_id")):
            continue
        add_em((e or {}).get("booker_email"))
        eid = (e or {}).get("id")
        for t in txns_by_eid.get(eid) or []:
            add_em((t or {}).get("booker_email"))
            add_em((t or {}).get("payer_email"))
    if not emails:
        return [], {}
    try:
        from pymongo.collation import Collation

        coll = Collation(locale="en", strength=2)
    except Exception:
        coll = None
    q = db.clients.find(
        {"email": {"$in": emails[:3000]}},
        {"_id": 0, "id": 1, "email": 1},
    )
    if coll is not None:
        q = q.collation(coll)
    rows = await q.to_list(len(emails) + 100)
    by_email: Dict[str, str] = {}
    out_ids: List[str] = []
    seen_id = set()
    for r in rows:
        cid = _norm_cmp_id(r.get("id"))
        em = (_clean_str(r.get("email")) or "").strip().lower()
        if not cid or not em:
            continue
        if em not in by_email:
            by_email[em] = cid
        if cid not in seen_id:
            seen_id.add(cid)
            out_ids.append(cid)
    return out_ids, by_email


async def _load_portal_clients_emi_context_map(client_ids: List[str]) -> Dict[str, dict]:
    """Minimal client projection for Home Coming EMI tier labels on admin participant reports."""
    uniq: List[str] = []
    seen = set()
    for x in client_ids or []:
        s = _norm_cmp_id(x)
        if s and s not in seen:
            seen.add(s)
            uniq.append(s)
    if not uniq:
        return {}
    rows = await db.clients.find(
        {"id": {"$in": uniq}},
        {"_id": 0, "id": 1, "subscription": 1, "annual_package_offer_prefs": 1},
    ).to_list(len(uniq) + 50)
    return {str(r.get("id") or "").strip(): r for r in rows if r.get("id") is not None}


async def _portal_cohort_by_client_ids(client_ids: List[str]) -> Dict[str, str]:
    """Map client id → portal cohort id (``awrp_batch_id``) for txns that store ``portal_client_id``."""
    uniq: List[str] = []
    seen = set()
    for x in client_ids or []:
        s = _norm_cmp_id(x)
        if s and s not in seen:
            seen.add(s)
            uniq.append(s)
    if not uniq:
        return {}
    rows = await db.clients.find(
        {"id": {"$in": uniq}},
        {"_id": 0, "id": 1, "awrp_batch_id": 1, "subscription": 1, "annual_subscription": 1},
    ).to_list(len(uniq) + 50)
    return {
        str(r.get("id") or "").strip(): _client_portal_cohort_from_client_doc(r)
        for r in rows
        if r.get("id") is not None
    }


async def _portal_cohort_by_contact_emails(
    enrollments: List[dict],
    txns_by_eid: Dict[str, List[dict]],
) -> Dict[str, str]:
    """Map normalized email → portal cohort id for booker, participants, and payment txn contacts."""
    uniq: List[str] = []
    seen = set()

    def add_em(raw: Any) -> None:
        em = (_clean_str(raw) or "").strip().lower()
        if em and em not in seen:
            seen.add(em)
            uniq.append(em)

    for e in enrollments:
        add_em(e.get("booker_email"))
        for p in (e.get("participants") or []):
            if isinstance(p, dict):
                add_em(_scalar_field(p, "email", "Email"))
        eid = e.get("id") or ""
        for t in txns_by_eid.get(eid) or []:
            add_em(t.get("booker_email"))
            add_em(t.get("payer_email"))
    if not uniq:
        return {}
    try:
        from pymongo.collation import Collation

        coll = Collation(locale="en", strength=2)
    except Exception:
        coll = None
    q = db.clients.find(
        {"email": {"$in": uniq}},
        {"_id": 0, "email": 1, "awrp_batch_id": 1, "subscription": 1, "annual_subscription": 1},
    )
    if coll is not None:
        q = q.collation(coll)
    rows = await q.to_list(20000)
    out: Dict[str, str] = {}
    for c in rows:
        ce = (_clean_str(c.get("email")) or "").strip().lower()
        bid = _client_portal_cohort_from_client_doc(c)
        if ce and bid and ce not in out:
            out[ce] = bid
    return out


async def _programs_duration_tiers_by_item_id(
    enrollments: List[dict],
    txns_by_eid: Dict[str, List[dict]],
) -> Dict[str, dict]:
    """Catalog program id → program doc subset (duration_tiers) for checkout tier/dates."""
    ids = set()
    for e in enrollments:
        eid = e.get("id") or ""
        txn = _pick_best_transaction_for_enrollment(e, txns_by_eid.get(eid) or [])
        it = (_clean_str(e.get("item_type")) or _clean_str((txn or {}).get("item_type"))).lower()
        if it != "program":
            continue
        e_pid = _clean_str(e.get("item_id"))
        if it == "program" and e_pid:
            pid = e_pid
        else:
            pid = _clean_str((txn or {}).get("item_id")) or e_pid
        if pid:
            ids.add(pid)
    if not ids:
        return {}
    id_list = list(ids)
    id_query: List[Any] = []
    for s in id_list:
        id_query.append(s)
        if s.isdigit():
            try:
                id_query.append(int(s))
            except (ValueError, TypeError):
                pass
    rows = await db.programs.find(
        {"id": {"$in": id_query}},
        {"_id": 0, "id": 1, "title": 1, "duration_tiers": 1},
    ).to_list(len(id_query) + 50)
    return {str(r.get("id")): r for r in rows if r.get("id") is not None}


def _tier_index_from_enrollment(enrollment: dict, txn: Optional[dict]) -> Optional[int]:
    raw = enrollment.get("tier_index")
    if raw is None and txn:
        raw = txn.get("tier_index")
    if raw is None or str(raw).strip() == "":
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _infer_tier_index_from_inr_list_prices(
    enrollment: dict,
    txn: Optional[dict],
    program: dict,
    seat_count: int,
) -> Optional[int]:
    """
    Legacy dashboard enrollments often omitted ``tier_index`` on the enrollment document.
    When checkout total is INR, pick the duration tier whose list ``price_inr`` is closest to the
    per-seat total, within a loose tolerance (discounts / GST).
    """
    if not isinstance(program, dict):
        return None
    tiers = program.get("duration_tiers") or []
    if not tiers:
        return None
    amt, cur = _payment_amount_currency(enrollment, txn)
    if txn and amt <= 0:
        try:
            alt = float(txn.get("enrollment_list_inr") or 0)
            if alt > 0:
                amt, cur = alt, "inr"
        except (TypeError, ValueError):
            pass
    if amt <= 0:
        try:
            dq = float(enrollment.get("dashboard_mixed_total") or 0)
        except (TypeError, ValueError):
            dq = 0.0
        if dq > 0 and _clean_str(enrollment.get("dashboard_mixed_currency")).lower() == "inr":
            amt, cur = dq, "inr"
    if amt <= 0 or (_clean_str(cur).lower() != "inr"):
        return None
    try:
        seats = max(1, int(seat_count or 1))
    except (TypeError, ValueError):
        seats = 1
    tx_pc = (txn or {}).get("participant_count")
    try:
        if tx_pc is not None and int(tx_pc) > 0:
            seats = max(seats, int(tx_pc))
    except (TypeError, ValueError):
        pass
    per_seat = float(amt) / float(seats)
    best_i: Optional[int] = None
    best_diff: Optional[float] = None
    for i, t in enumerate(tiers):
        if not isinstance(t, dict):
            continue
        try:
            list_inr = float(t.get("price_inr") or 0)
        except (TypeError, ValueError):
            list_inr = 0.0
        try:
            offer_inr = float(t.get("offer_price_inr") or 0)
        except (TypeError, ValueError):
            offer_inr = 0.0
        anchors = [x for x in (list_inr, offer_inr) if x > 0]
        if not anchors:
            continue
        anchor = min(anchors, key=lambda x: abs(per_seat - x))
        diff = abs(per_seat - anchor)
        if best_diff is None or diff < best_diff:
            best_diff = diff
            best_i = i
    if best_i is None or best_diff is None:
        return None
    tier_at = tiers[best_i] if isinstance(tiers[best_i], dict) else {}
    try:
        list_ref = float(tier_at.get("price_inr") or 0)
        off_ref = float(tier_at.get("offer_price_inr") or 0)
        ref = max([x for x in (list_ref, off_ref) if x > 0] or [0.0])
    except (TypeError, ValueError):
        return None
    if ref <= 0:
        return None
    # Loose match: India totals include GST/platform; cohort / CRM discounts widen the gap.
    if best_diff / ref <= 0.55 or best_diff <= 12000:
        return int(best_i)
    return None


def _is_three_month_duration_tier(tier: Optional[dict]) -> bool:
    if not isinstance(tier, dict):
        return False
    label = (tier.get("label") or "").lower()
    if re.search(r"\b3\s*[- ]?\s*months?\b", label):
        return True
    if re.search(r"\bthree\s+months?\b", label):
        return True
    dm = tier.get("duration_months")
    try:
        if int(dm) == 3:
            return True
    except (TypeError, ValueError):
        pass
    dur = (tier.get("duration") or "").lower()
    if "3 month" in dur or "90 day" in dur or "90 days" in dur:
        return True
    return False


def _program_tier_fields_for_report(
    item_type: str,
    catalog_program_id: str,
    tier_index: Optional[int],
    programs_by_id: Optional[Dict[str, dict]],
) -> Dict[str, Any]:
    """Chosen tier label + tier window from catalog ``duration_tiers`` (program checkouts only)."""
    base: Dict[str, Any] = {
        "program_catalog_id": _clean_str(catalog_program_id),
        "catalog_program_title": "",
        "tier_index": None,
        "tier_label": "",
        "chosen_start_date": "",
        "chosen_end_date": "",
        "is_three_month_tier": False,
    }
    it = (item_type or "").lower()
    if it != "program" or not catalog_program_id or not programs_by_id:
        return base
    prog = programs_by_id.get(catalog_program_id)
    if isinstance(prog, dict):
        base["catalog_program_title"] = _clean_str(prog.get("title"))
    if tier_index is None:
        return base
    if not prog:
        base["tier_index"] = tier_index
        return base
    tiers = prog.get("duration_tiers") or []
    if tier_index < 0 or tier_index >= len(tiers):
        base["tier_index"] = tier_index
        return base
    tier = tiers[tier_index] if isinstance(tiers[tier_index], dict) else {}
    base["tier_index"] = tier_index
    base["tier_label"] = _clean_str(tier.get("label")) or f"Tier {tier_index + 1}"
    base["chosen_start_date"] = _clean_str(tier.get("start_date"))
    base["chosen_end_date"] = _clean_str(tier.get("end_date"))
    base["is_three_month_tier"] = bool(_is_three_month_duration_tier(tier))
    return base


def _annual_emi_cadence_word(prefs: dict, subscription: dict) -> str:
    """Human cadence for EMI tier column: monthly / quarterly / yearly."""
    pm = str((prefs or {}).get("payment_mode") or "").strip().lower()
    if pm == "emi_monthly":
        return "monthly"
    if pm == "emi_quarterly":
        return "quarterly"
    if pm == "emi_yearly":
        return "yearly"
    if str((subscription or {}).get("payment_mode") or "").strip().upper() == "EMI":
        emis = (subscription or {}).get("emis") or []
        n = len(emis) if isinstance(emis, list) else 0
        if n >= 12:
            return "monthly"
        if n > 4:
            return "quarterly"
        if n > 1:
            return "yearly"
    return ""


def _txn_inr_amount_tolerance(tx_amt: float) -> float:
    """Allow GST / platform drift between checkout txn amount and CRM EMI split rows."""
    if tx_amt <= 0:
        return 2.0
    return max(75.0, min(float(tx_amt) * 0.12, 15000.0))


def _emi_row_amount(emi: dict) -> float:
    for k in ("amount", "amount_inr", "Amount"):
        try:
            v = float(emi.get(k) or 0)
            if v > 0:
                return v
        except (TypeError, ValueError):
            continue
    return 0.0


def _emi_installment_match_for_txn(txn: dict, emis: List[dict]) -> Optional[Tuple[int, int]]:
    """
    Map a paid checkout txn to (installment_number, total_installments) using CRM EMI rows.
    Prefers transaction_id vs session / Razorpay ids; then amount match (paid rows, then due/pending).
    """
    if not txn or not emis:
        return None
    st_pay = str(txn.get("payment_status") or "").lower()
    if st_pay not in ("paid", "complete", "completed"):
        return None
    total = len(emis)
    if total < 2:
        return None
    sid = str(txn.get("stripe_session_id") or "").strip()
    rzp = str(txn.get("razorpay_payment_id") or "").strip()
    rz_order = str(txn.get("razorpay_order_id") or "").strip()
    try:
        tx_amt = float(txn.get("amount") or 0)
    except (TypeError, ValueError):
        tx_amt = 0.0
    tol = _txn_inr_amount_tolerance(tx_amt)

    def ref_hit(emi: dict) -> bool:
        tid = str(emi.get("transaction_id") or "").strip()
        if not tid:
            return False
        if sid and (tid == sid or sid in tid or tid in sid):
            return True
        if rzp and (tid == rzp or rzp in tid or tid in rzp):
            return True
        if rz_order and (tid == rz_order or rz_order in tid or tid in rz_order):
            return True
        return False

    ref_ns: List[int] = []
    for emi in emis:
        if not isinstance(emi, dict):
            continue
        if ref_hit(emi):
            try:
                n = int(emi.get("number") or emi.get("emi_number") or 0)
            except (TypeError, ValueError):
                continue
            if n >= 1:
                ref_ns.append(n)
    if ref_ns:
        return min(ref_ns), total

    def amount_match_pool(paid_only: bool) -> List[dict]:
        out: List[dict] = []
        for x in emis:
            if not isinstance(x, dict):
                continue
            s = str(x.get("status") or "").strip().lower()
            if paid_only:
                if s == "paid":
                    out.append(x)
            else:
                if s not in ("rejected", "cancelled", "canceled"):
                    out.append(x)
        return out

    for paid_only in (True, False):
        pool = amount_match_pool(paid_only)
        cand: List[dict] = []
        for emi in pool:
            ea = _emi_row_amount(emi)
            if ea > 0 and tx_amt > 0 and abs(ea - tx_amt) <= tol:
                cand.append(emi)
        if not cand:
            continue
        cand.sort(key=_emi_sort_number)
        try:
            n0 = int(cand[0].get("number") or cand[0].get("emi_number") or 0)
        except (TypeError, ValueError):
            continue
        if n0 >= 1:
            return n0, total
    return None


def _maybe_overlay_home_coming_emi_tier_label(
    tier_fields: dict,
    *,
    enrollment: dict,
    program_display: str,
    txn_for_amount: Optional[dict],
    all_txns: List[dict],
    clients_by_id: Optional[Dict[str, dict]],
    portal_client_id: str,
) -> None:
    """
    Replace misleading catalog tier (e.g. '1 Month') when the linked CRM client has a multi-installment
    subscription schedule: prefer 'Annual Program — EMI k/n — cadence'; if the installment cannot be
    matched to a txn, still set 'Annual program · EMI (n installments)' so admins do not read EMI as a
    one-month program.
    """
    if not clients_by_id or not tier_fields:
        return
    pcid = _norm_cmp_id(portal_client_id)
    ecid = _norm_cmp_id((enrollment or {}).get("client_id"))
    cl = None
    if pcid:
        cl = clients_by_id.get(pcid)
    if cl is None and ecid:
        cl = clients_by_id.get(ecid)
    if not isinstance(cl, dict):
        return
    sub = cl.get("subscription") if isinstance(cl.get("subscription"), dict) else {}
    emis = sub.get("emis")
    if not isinstance(emis, list):
        return
    n_sched = sum(1 for x in emis if isinstance(x, dict))
    if n_sched < 2:
        return
    prefs = cl.get("annual_package_offer_prefs") if isinstance(cl.get("annual_package_offer_prefs"), dict) else {}
    pm_pref = str(prefs.get("payment_mode") or "").strip().lower()
    pm_sub = str(sub.get("payment_mode") or "").strip().upper()
    try:
        num_emis_sub = int(sub.get("num_emis") or 0)
    except (TypeError, ValueError):
        num_emis_sub = 0
    looks_emi = (
        pm_sub == "EMI"
        or pm_pref.startswith("emi_")
        or num_emis_sub >= 2
        or n_sched >= 3
    )
    if not looks_emi:
        return

    tier_now = _clean_str(tier_fields.get("tier_label"))
    if not _eligible_for_annual_emi_admin_tier_tag(
        enrollment or {},
        program_display,
        tier_now,
        emi_row_count=n_sched,
    ):
        return

    ordered_txns = sorted(
        [t for t in (all_txns or []) if str(t.get("payment_status") or "").lower() in ("paid", "complete", "completed")],
        key=_txn_sort_key,
        reverse=True,
    )
    hit: Optional[Tuple[int, int]] = None
    for t in ordered_txns:
        hit = _emi_installment_match_for_txn(t, emis)
        if hit:
            break
    if not hit and txn_for_amount:
        hit = _emi_installment_match_for_txn(txn_for_amount, emis)
    if not hit:
        tier_fields["tier_label"] = f"Annual program · EMI ({n_sched} installments)"
        return
    emi_n, total_n = hit
    if emi_n <= 0:
        tier_fields["tier_label"] = f"Annual program · EMI ({n_sched} installments)"
        return
    cadence = _annual_emi_cadence_word(prefs, sub)
    # e.g. "Annual Program — EMI 1/12 — monthly" or "… — EMI 12/12 (final) — monthly"
    emi_slot = f"EMI {emi_n}/{total_n}"
    if total_n > 1 and emi_n >= total_n:
        emi_slot = f"EMI {emi_n}/{total_n} (final)"
    bits = ["Annual Program", emi_slot]
    if cadence:
        bits.append(cadence)
    tier_fields["tier_label"] = " — ".join(bits)


def build_participant_report_rows(
    enrollments: List[dict],
    txns_by_eid: Dict[str, List[dict]],
    *,
    paid_completed_only: bool = False,
    programs_by_id: Optional[Dict[str, dict]] = None,
    portal_cohort_by_email: Optional[Dict[str, str]] = None,
    portal_cohort_by_client_id: Optional[Dict[str, str]] = None,
    portal_clients_by_id: Optional[Dict[str, dict]] = None,
    booker_email_to_client_id: Optional[Dict[str, str]] = None,
) -> List[dict]:
    """
    One row per participant (or one booker row if participants missing).
    Includes enrollment form fields: relationship, gender, city, state, attendance, notify, referral, etc.
    Covers dashboard, cart, program page, session page, upcoming — all use `enrollments`.
    """
    rows: List[dict] = []
    done_status = frozenset(
        {
            "completed",
            "paid",
            "india_payment_approved",
        }
    )

    for e in enrollments:
        eid = e.get("id") or ""
        st = (e.get("status") or "").lower()
        txn = _pick_best_transaction_for_enrollment(e, txns_by_eid.get(eid) or [])
        pay_st = _clean_str((txn or {}).get("payment_status")).lower()
        is_done = (
            pay_st in ("paid", "complete", "completed")
            or st in done_status
            or e.get("step") == 5
        )

        if paid_completed_only and not is_done:
            continue

        amt, cur = _payment_amount_currency(e, txn)
        inv = _clean_str((txn or {}).get("invoice_number")) or _clean_str(e.get("invoice_number"))
        participants = _merge_participants_for_report(e, txn)
        program = display_program_title_for_enrollment(e, participants, None)
        item_type = _clean_str(e.get("item_type")) or _clean_str((txn or {}).get("item_type"))
        booker_name = _clean_str(e.get("booker_name"))
        booker_email = _clean_str(e.get("booker_email"))
        booker_phone = _clean_str(e.get("phone"))
        booker_country = _clean_str(e.get("booker_country"))
        created = _clean_str(e.get("created_at"))
        origin = _enrollment_origin(e)

        booker_email_key = (_clean_str(e.get("booker_email")) or "").strip().lower()
        payer_email_key = (_clean_str((txn or {}).get("payer_email")) or "").strip().lower()
        txn_booker_email_key = (_clean_str((txn or {}).get("booker_email")) or "").strip().lower()
        portal_cohort = ""
        cid_key = _norm_cmp_id((txn or {}).get("portal_client_id"))
        if cid_key and portal_cohort_by_client_id:
            portal_cohort = (portal_cohort_by_client_id or {}).get(cid_key, "") or ""
        if portal_cohort_by_email:
            portal_cohort = portal_cohort or portal_cohort_by_email.get(booker_email_key, "")
            portal_cohort = portal_cohort or portal_cohort_by_email.get(payer_email_key, "")
            portal_cohort = portal_cohort or portal_cohort_by_email.get(txn_booker_email_key, "")
        if item_type.lower() == "program" and _clean_str(e.get("item_id")):
            catalog_pid = _clean_str(e.get("item_id"))
        else:
            catalog_pid = _clean_str((txn or {}).get("item_id")) or _clean_str(e.get("item_id"))
        total_slots = len(participants) if participants else 1
        tier_idx = _tier_index_from_enrollment(e, txn)
        if tier_idx is None and item_type.lower() == "program" and catalog_pid and programs_by_id:
            prog_doc = programs_by_id.get(catalog_pid)
            if isinstance(prog_doc, dict):
                inferred = _infer_tier_index_from_inr_list_prices(e, txn, prog_doc, total_slots)
                if inferred is not None:
                    tier_idx = inferred
        tier_fields = _program_tier_fields_for_report(item_type, catalog_pid, tier_idx, programs_by_id)
        p_cs = _clean_str(e.get("chosen_start_date"))
        p_ce = _clean_str(e.get("chosen_end_date"))
        p_cl = _clean_str(e.get("chosen_tier_label"))
        if p_cs:
            tier_fields["chosen_start_date"] = p_cs
        if p_ce:
            tier_fields["chosen_end_date"] = p_ce
        if p_cl:
            tier_fields["tier_label"] = p_cl
        tier_fields["portal_cohort"] = portal_cohort
        pcid_emi = cid_key or _norm_cmp_id(e.get("client_id")) or ""
        txn_list = txns_by_eid.get(eid) or []
        if not pcid_emi and txn_list:
            for t in txn_list:
                pcid_emi = _norm_cmp_id(t.get("portal_client_id"))
                if pcid_emi:
                    break
        if not pcid_emi and booker_email_to_client_id:
            for em_key in (booker_email_key, payer_email_key, txn_booker_email_key):
                pcid_emi = _norm_cmp_id((booker_email_to_client_id or {}).get(em_key, ""))
                if pcid_emi:
                    break
        _maybe_overlay_home_coming_emi_tier_label(
            tier_fields,
            enrollment=e,
            program_display=program,
            txn_for_amount=txn,
            all_txns=txn_list,
            clients_by_id=portal_clients_by_id,
            portal_client_id=pcid_emi,
        )
        tl_after = _clean_str(tier_fields.get("tier_label"))
        if (
            _tier_label_implies_catalog_one_month(tl_after)
            and _enrollment_looks_like_home_coming(e, program)
            and "annual program" not in tl_after.lower()
        ):
            tier_fields["tier_label"] = "Annual program · EMI"

        def push_row(p: Optional[dict], index: int, *, booker_only: bool) -> None:
            if booker_only or p is None:
                ph = booker_phone
                wa = booker_phone
                row = {
                    "invoice_number": inv,
                    "enrollment_id": eid,
                    "program": program,
                    "item_type": item_type,
                    "enrollment_status": _clean_str(e.get("status")),
                    "enrollment_origin": origin,
                    "booker_name": booker_name,
                    "booker_email": booker_email,
                    "booker_phone": booker_phone,
                    "booker_country": booker_country,
                    "participant_index": 1,
                    "participant_total": 1,
                    "participant_name": booker_name,
                    "relationship": "",
                    "age": "",
                    "gender": "",
                    "country": booker_country,
                    "city": "",
                    "state": "",
                    "attendance_mode": "",
                    "notify_enrollment": "",
                    "participant_email": booker_email,
                    "phone": ph,
                    "whatsapp": wa,
                    "is_first_time": "",
                    "referral_source": "",
                    "referred_by_name": "",
                    "referred_by_email": "",
                    "participant_program_id": "",
                    "participant_program_title": "",
                    "participant_uid": "",
                    "payment_amount": amt,
                    "payment_currency": cur,
                    "payment_status": pay_st or _clean_str(e.get("status")),
                    "created_at": created,
                }
                row.update(tier_fields)
                rows.append(row)
                return

            p_phone = _clean_str(_scalar_field(p, "phone", "Phone"))
            p_wa = _clean_str(_scalar_field(p, "whatsapp", "WhatsApp"))
            ph = p_phone or booker_phone
            wa = (p_wa or p_phone) or ph
            notify_yes = _notify_yes_from_participant(p)
            first_time_v = _scalar_field(p, "is_first_time", "isFirstTime")
            is_first = bool(first_time_v) if isinstance(first_time_v, bool) else str(first_time_v).strip().lower() in ("1", "true", "yes")
            row = {
                "invoice_number": inv,
                "enrollment_id": eid,
                "program": program,
                "item_type": item_type,
                "enrollment_status": _clean_str(e.get("status")),
                "enrollment_origin": origin,
                "booker_name": booker_name,
                "booker_email": booker_email,
                "booker_phone": booker_phone,
                "booker_country": booker_country,
                "participant_index": index,
                "participant_total": total_slots,
                "participant_name": _clean_str(_scalar_field(p, "name", "Name")) or booker_name,
                "relationship": _clean_str(_scalar_field(p, "relationship", "Relationship")),
                "age": _format_age_for_report(_scalar_field(p, "age", "Age")),
                "gender": _clean_str(_scalar_field(p, "gender", "Gender")),
                "country": _clean_str(_scalar_field(p, "country", "Country")) or booker_country,
                "city": _clean_str(_scalar_field(p, "city", "City")),
                "state": _clean_str(_scalar_field(p, "state", "State")),
                "attendance_mode": _attendance_mode_from_participant(p),
                "notify_enrollment": "Yes" if notify_yes else "No",
                "participant_email": _clean_str(_scalar_field(p, "email", "Email")),
                "phone": ph,
                "whatsapp": wa,
                "is_first_time": "Yes" if is_first else "No",
                "referral_source": _clean_str(_scalar_field(p, "referral_source", "referralSource")),
                "referred_by_name": _clean_str(_scalar_field(p, "referred_by_name", "referredByName")),
                "referred_by_email": _clean_str(_scalar_field(p, "referred_by_email", "referredByEmail")),
                "participant_program_id": _clean_str(_scalar_field(p, "program_id", "programId")),
                "participant_program_title": _clean_str(_scalar_field(p, "program_title", "programTitle")),
                "participant_uid": _clean_str(_scalar_field(p, "uid", "UID")),
                "payment_amount": amt,
                "payment_currency": cur,
                "payment_status": pay_st or _clean_str(e.get("status")),
                "created_at": created,
            }
            row.update(tier_fields)
            rows.append(row)

        if not participants:
            push_row(None, 1, booker_only=True)
        else:
            for i, p in enumerate(participants, start=1):
                push_row(p, i, booker_only=False)

    return rows


UPLOAD_DIR = ROOT_DIR / "uploads" / "payment_proofs"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


@router.post("/submit-proof")
async def submit_payment_proof(
    enrollment_id: str = Form(...),
    payer_name: str = Form(...),
    payer_email: str = Form(""),
    payer_phone: str = Form(""),
    payment_date: str = Form(...),
    bank_name: str = Form(""),
    transaction_id: str = Form(...),
    amount: str = Form(...),
    city: str = Form(""),
    state: str = Form(""),
    payment_method: str = Form(""),
    program_type: str = Form(""),
    selected_item: str = Form(""),
    is_emi: str = Form("false"),
    emi_total_months: str = Form(""),
    emi_months_covered: str = Form(""),
    notes: str = Form(""),
    screenshot: Optional[UploadFile] = File(None),
):
    """Submit India alternative payment proof for admin approval."""
    # Validate enrollment exists (skip for standalone 'MANUAL' submissions)
    enrollment = {}
    if enrollment_id != "MANUAL":
        enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0}) or {}

    pm = (payment_method or "").strip().lower()
    requires_screenshot = pm in {"cash_deposit", "cheque"}
    if requires_screenshot:
        if screenshot is None or not (getattr(screenshot, "filename", None) or "").strip():
            if pm == "cheque":
                detail = "Payment screenshot is required for cheque (e.g. scan or photo of the cheque / deposit proof)."
            else:
                detail = "Payment screenshot is required for cash deposit (e.g. deposit slip or receipt)."
            raise HTTPException(status_code=400, detail=detail)

    if pm in {"cash_deposit", "cheque"} and not (notes or "").strip():
        raise HTTPException(
            status_code=400,
            detail="Notes are required: who deposited or issued the cheque, which city, and which bank / branch.",
        )

    if pm == "upi" and not (notes or "").strip():
        raise HTTPException(
            status_code=400,
            detail="Notes are required for UPI / GPay: who paid (name as on the app) and when you paid (date/time if possible).",
        )

    screenshot_public_url = ""
    if screenshot is not None and (getattr(screenshot, "filename", None) or "").strip():
        # Save screenshot (S3 when configured, else local disk unless REQUIRE_S3_FOR_UPLOADS)
        ext = screenshot.filename.split(".")[-1] if "." in screenshot.filename else "png"
        filename = f"{uuid.uuid4().hex[:12]}.{ext}"
        proof_bytes = await screenshot.read()
        mime, _ = mimetypes.guess_type(filename)
        mime = mime or "image/png"
        must_s3 = s3_storage.media_must_use_s3()
        if must_s3 and not s3_storage.is_s3_enabled():
            raise HTTPException(
                status_code=503,
                detail=(
                    "Payment proof screenshots must be stored in S3 (REQUIRE_S3_FOR_UPLOADS) but S3 is not configured. "
                    "See GET /api/upload/storage-status."
                ),
            )
        if s3_storage.is_s3_enabled():
            key = s3_storage.payment_proof_key(filename)
            try:
                screenshot_public_url = s3_storage.upload_bytes(key, proof_bytes, mime)
            except Exception as e:
                if must_s3:
                    raise HTTPException(status_code=503, detail=f"S3 upload failed: {e}") from e
                logger.warning("S3 payment proof upload failed; saving locally: %s", e)
                filepath = UPLOAD_DIR / filename
                with open(filepath, "wb") as f:
                    f.write(proof_bytes)
                screenshot_public_url = f"/api/uploads/payment_proofs/{filename}"
        else:
            filepath = UPLOAD_DIR / filename
            with open(filepath, "wb") as f:
                f.write(proof_bytes)
            screenshot_public_url = f"/api/uploads/payment_proofs/{filename}"

    proof = {
        "id": str(uuid.uuid4()),
        "enrollment_id": enrollment_id,
        "booker_name": enrollment.get("booker_name", payer_name),
        "booker_email": enrollment.get("booker_email", payer_email),
        "payer_name": payer_name,
        "payer_email": payer_email,
        "payer_phone": payer_phone,
        "payment_date": payment_date,
        "bank_name": bank_name,
        "transaction_id": transaction_id,
        "program_type": program_type,
        "selected_item": selected_item,
        "item_id": (enrollment.get("item_id") or "").strip() if enrollment else "",
        "item_type": ((enrollment.get("item_type") or "program") if enrollment else "program").strip().lower(),
        "program_title": enrollment.get("item_title", selected_item or program_type),
        "amount": amount,
        "city": city,
        "state": state,
        "payment_method": payment_method,
        "is_emi": is_emi == "true",
        "emi_total_months": int(emi_total_months) if emi_total_months else None,
        "emi_months_covered": int(emi_months_covered) if emi_months_covered else None,
        "notes": notes,
        "screenshot_url": screenshot_public_url,
        "status": "pending",
        "participants": enrollment.get("participants", []),
        "participant_count": enrollment.get("participant_count", 1),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.india_payment_proofs.insert_one(proof)

    # Update enrollment status if linked
    if enrollment_id != "MANUAL" and enrollment:
        await db.enrollments.update_one(
            {"id": enrollment_id},
            {"$set": {
                "status": "india_payment_proof_submitted",
                "india_payment_proof_id": proof["id"],
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )

    logger.info(f"[INDIA PAYMENT PROOF] enrollment={enrollment_id}, txn={transaction_id}, amount={amount}")
    return {"message": "Payment proof submitted successfully. Awaiting admin approval.", "proof_id": proof["id"]}


@router.get("/admin/list")
async def list_payment_proofs():
    """Admin: list all India payment proofs."""
    proofs = await db.india_payment_proofs.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return proofs


@router.put("/admin/proofs/{proof_id}")
async def update_payment_proof(proof_id: str, data: dict):
    """Admin: edit a submitted payment proof before approving."""
    allowed_fields = ['payer_name', 'booker_email', 'amount', 'transaction_id', 'program_title',
                       'bank_name', 'payment_date', 'payment_method', 'city', 'state', 'admin_notes', 'phone']
    update = {k: v for k, v in data.items() if k in allowed_fields and v is not None}
    if 'amount' in update:
        try: update['amount'] = float(update['amount'])
        except: pass
    update['updated_at'] = datetime.now(timezone.utc).isoformat()
    await db.india_payment_proofs.update_one({"id": proof_id}, {"$set": update})
    return {"message": "Proof updated"}


def _titles_blob_for_home_coming_check(proof: dict, enrollment: dict) -> str:
    parts = [
        proof.get("program_title"),
        proof.get("selected_item"),
        proof.get("program_type"),
        (enrollment or {}).get("item_title"),
    ]
    return " ".join(str(p or "") for p in parts).lower()


def _is_home_coming_context(proof: dict, enrollment: dict) -> bool:
    blob = _titles_blob_for_home_coming_check(proof, enrollment)
    compact = blob.replace(" ", "")
    return "home coming" in blob or "homecoming" in compact


def _enrollment_looks_like_home_coming(enrollment: dict, program_display: str = "") -> bool:
    """True when reporting row is clearly Home Coming (participant titles, item title, resolved program label)."""
    if not enrollment:
        return False
    parts = [
        enrollment.get("item_title"),
        enrollment.get("program_title"),
        enrollment.get("selected_program_title"),
        program_display,
    ]
    for p in enrollment.get("participants") or []:
        if isinstance(p, dict):
            parts.append(_scalar_field(p, "program_title", "programTitle"))
    for line in enrollment.get("portal_cart_lines") or []:
        if isinstance(line, dict):
            parts.append(line.get("title") or line.get("program_title") or line.get("name"))
    blob = " ".join(str(p or "") for p in parts).lower()
    compact = blob.replace(" ", "")
    return "home coming" in blob or "homecoming" in compact


def _tier_label_implies_catalog_one_month(label: str) -> bool:
    """Catalog / chosen tier text that reads like a single-month option (misleading on EMI annual)."""
    s = (label or "").strip().lower()
    if not s:
        return False
    if re.search(r"\b1\s*[-]?\s*months?\b", s):
        return True
    if re.search(r"\bone\s+months?\b", s):
        return True
    if re.search(r"\b30\s*[-]?\s*days?\b", s):
        return True
    return False


def _eligible_for_annual_emi_admin_tier_tag(
    enrollment: dict,
    program_display: str,
    tier_label: str,
    *,
    emi_row_count: int,
) -> bool:
    """Whether we should replace or tag tier text so admins do not read EMI annual as a one-month program."""
    if emi_row_count < 2:
        return False
    if _enrollment_looks_like_home_coming(enrollment or {}, program_display):
        return True
    # Dashboard / cart rows: catalog may still say "1 Month" while CRM has a multi-installment schedule.
    if emi_row_count >= 3 and _tier_label_implies_catalog_one_month(tier_label):
        return True
    return False


def _emi_sort_number(emi: dict) -> int:
    try:
        return int(emi.get("number") or emi.get("emi_number") or 999)
    except (TypeError, ValueError):
        return 999


def _mark_subscription_emi_paid_from_india_proof(emi: dict, proof: dict) -> bool:
    """Update one EMI dict in-place to match an approved India manual proof. Returns True if changed."""
    if not isinstance(emi, dict):
        return False
    if str(emi.get("status", "")).lower() == "paid":
        return False
    pay_raw = str(proof.get("payment_date") or "").strip()
    pay_date = pay_raw[:10] if len(pay_raw) >= 10 else ""
    if not pay_date:
        pay_date = datetime.now(timezone.utc).date().isoformat()
    pm = (proof.get("payment_method") or "").strip().lower() or "manual_proof"
    emi["status"] = "paid"
    emi["date"] = pay_date
    emi["paid_date"] = pay_date
    emi["payment_method"] = pm
    emi["transaction_id"] = str(proof.get("transaction_id") or "").strip()
    emi["paid_by"] = str(
        proof.get("payer_name") or proof.get("booker_name") or ""
    ).strip()
    su = str(proof.get("screenshot_url") or "").strip()
    rec = str(proof.get("receipt_url") or "").strip()
    if su:
        emi["receipt_url"] = su
    elif rec:
        emi["receipt_url"] = rec
    emi["remaining"] = 0.0
    pid = proof.get("id")
    if pid:
        emi["india_payment_proof_id"] = pid
    return True


async def _apply_home_coming_emi_payment_sync(
    *,
    client_id: str,
    proof: dict,
    enrollment: dict,
    log_prefix: str = "Home Coming EMIs synced",
    log_extra_id: Any = None,
) -> None:
    """
    Student Home Coming schedule reads ``clients.subscription.emis`` from GET /student/home.
    Used for India proof approval and for Razorpay/Stripe enrollment payments.
    """
    cid = (client_id or "").strip()
    if not cid or not _is_home_coming_context(proof, enrollment or {}):
        return
    cl = await db.clients.find_one({"id": cid})
    if not cl:
        return
    sub = dict(cl.get("subscription") or {})
    emis = sub.get("emis")
    if not isinstance(emis, list) or len(emis) == 0:
        return
    emis = [dict(e) if isinstance(e, dict) else e for e in emis]

    try:
        amt = float(proof.get("amount") or 0)
    except (TypeError, ValueError):
        amt = 0.0

    changed = False
    if proof.get("is_emi"):
        try:
            n = int(proof.get("emi_months_covered") or 0)
        except (TypeError, ValueError):
            n = 0
        if n >= 1:
            for emi in emis:
                if not isinstance(emi, dict):
                    continue
                try:
                    en = int(emi.get("number") or emi.get("emi_number") or 0)
                except (TypeError, ValueError):
                    en = 0
                if en == n:
                    changed = _mark_subscription_emi_paid_from_india_proof(emi, proof) or changed
                    break
    else:
        unpaid = [
            e for e in emis if isinstance(e, dict) and str(e.get("status", "")).lower() != "paid"
        ]
        if unpaid:
            total_unpaid = 0.0
            for e in unpaid:
                try:
                    total_unpaid += float(e.get("amount") or 0)
                except (TypeError, ValueError):
                    pass
            tol = max(2.0, total_unpaid * 0.02) if total_unpaid > 0 else 2.0
            if total_unpaid > 0 and amt + 1e-6 >= total_unpaid - tol:
                for emi in unpaid:
                    changed = _mark_subscription_emi_paid_from_india_proof(emi, proof) or changed
            # One installment via Razorpay/Stripe (no emi_months_covered) — match amount to a row.
            if not changed and amt > 0:
                tol_emi = max(2.0, amt * 0.02)
                matches: List[dict] = []
                for e in unpaid:
                    if not isinstance(e, dict) or str(e.get("status", "")).lower() == "paid":
                        continue
                    try:
                        ea = float(e.get("amount") or 0)
                    except (TypeError, ValueError):
                        continue
                    if ea > 0 and abs(ea - amt) <= tol_emi:
                        matches.append(e)
                if matches:
                    pick = sorted(matches, key=_emi_sort_number)[0]
                    changed = _mark_subscription_emi_paid_from_india_proof(pick, proof) or changed

    if not changed:
        return
    sub["emis"] = emis
    await db.clients.update_one(
        {"id": cid},
        {
            "$set": {
                "subscription": sub,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
        },
    )
    logger.info("%s client=%s ref=%s", log_prefix, cid, log_extra_id or proof.get("id"))


def _proof_like_from_enrollment_transaction(enrollment: dict, transaction: dict) -> dict:
    pay_raw = ""
    pa = transaction.get("paid_at")
    if isinstance(pa, str) and pa.strip():
        pay_raw = pa.strip()
    elif hasattr(pa, "isoformat"):
        try:
            pay_raw = pa.isoformat()
        except Exception:
            pay_raw = ""
    if not pay_raw:
        ua = transaction.get("updated_at")
        if isinstance(ua, str) and ua.strip():
            pay_raw = ua.strip()
        elif hasattr(ua, "isoformat"):
            try:
                pay_raw = ua.isoformat()
            except Exception:
                pay_raw = ""
    try:
        amt = float(transaction.get("amount") or 0)
    except (TypeError, ValueError):
        amt = 0.0
    pm_from_tx = str(transaction.get("payment_method") or "").strip().lower()
    prov = str(transaction.get("payment_provider") or "").strip().lower()
    if prov == "razorpay":
        pm = "razorpay"
    elif pm_from_tx == "manual_proof":
        pm = "manual_proof"
    elif prov == "stripe":
        pm = "stripe"
    elif prov:
        pm = prov
    elif pm_from_tx:
        pm = pm_from_tx
    else:
        pm = str(enrollment.get("payment_method") or "stripe").lower()
    tid = str(transaction.get("razorpay_payment_id") or "").strip()
    if not tid:
        tid = str(transaction.get("stripe_session_id") or "").strip()
    sid = str(transaction.get("stripe_session_id") or "").strip()
    receipt_url = f"/payment/success?session_id={sid}" if sid else ""
    return {
        "program_title": transaction.get("item_title") or enrollment.get("item_title"),
        "selected_item": enrollment.get("item_id") or transaction.get("item_id"),
        "program_type": str(
            enrollment.get("program_type") or enrollment.get("item_type") or transaction.get("item_type") or ""
        ),
        "amount": amt,
        "payment_method": pm,
        "transaction_id": tid,
        "payment_date": pay_raw,
        "is_emi": bool(enrollment.get("checkout_is_emi")),
        "emi_months_covered": enrollment.get("checkout_emi_months_covered"),
        "payer_name": transaction.get("booker_name"),
        "booker_name": transaction.get("booker_name"),
        "screenshot_url": "",
        "receipt_url": receipt_url,
        "id": None,
    }


def _merge_india_proof_into_proof_like(proof_like: dict, proof_doc: dict, tx: dict) -> dict:
    """Enrich synthetic proof_like with fields stored only on india_payment_proofs (EMI month, UPI proof, etc.)."""
    out = dict(proof_like)
    if (proof_doc.get("status") or "").strip().lower() != "approved":
        return out
    if proof_doc.get("is_emi") is not None:
        out["is_emi"] = bool(proof_doc.get("is_emi"))
    if proof_doc.get("emi_months_covered") is not None:
        out["emi_months_covered"] = proof_doc.get("emi_months_covered")
    for k in ("program_title", "selected_item", "program_type"):
        v = proof_doc.get(k)
        if isinstance(v, str) and v.strip():
            out[k] = v.strip()
    pd = proof_doc.get("payment_date")
    if isinstance(pd, str) and len(pd.strip()) >= 8:
        out["payment_date"] = pd.strip()
    ptid = str(proof_doc.get("transaction_id") or "").strip()
    if ptid:
        out["transaction_id"] = ptid
    su = str(proof_doc.get("screenshot_url") or "").strip()
    if su:
        out["screenshot_url"] = su
    out["payer_name"] = proof_doc.get("payer_name") or out.get("payer_name")
    out["booker_name"] = proof_doc.get("booker_name") or out.get("booker_name")
    out["id"] = proof_doc.get("id")
    try:
        out["amount"] = float(tx.get("amount") if tx.get("amount") is not None else proof_doc.get("amount") or 0)
    except (TypeError, ValueError):
        pass
    pm = (proof_doc.get("payment_method") or "").strip().lower()
    if pm:
        out["payment_method"] = pm
    return out


async def _resolve_portal_client_id_for_home_coming_emi(tx: dict, enrollment: dict) -> str:
    pcid = (tx.get("portal_client_id") or "").strip()
    if pcid:
        return pcid
    ce = str(enrollment.get("client_id") or "").strip()
    if ce:
        return ce
    be = (enrollment.get("booker_email") or tx.get("booker_email") or "").strip().lower()
    if be:
        crow = await db.clients.find_one({"email": be}, {"id": 1})
        if crow:
            return str(crow.get("id") or "").strip()
    return ""


async def sync_home_coming_emis_for_paid_enrollment_tx(tx: dict) -> None:
    """
    Update Home Coming ``clients.subscription.emis`` from a paid enrollment transaction.
    Called from the receipt pipeline so manual India approvals stay in sync as soon as receipts go out,
    and from Razorpay/Stripe hooks when no receipt is sent.
    """
    if not tx or str(tx.get("payment_status") or "").lower() != "paid":
        return
    enrollment_id = (tx.get("enrollment_id") or "").strip()
    if not enrollment_id:
        return
    enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0})
    if not enrollment:
        return
    pcid = await _resolve_portal_client_id_for_home_coming_emi(tx, enrollment)
    if not pcid:
        return
    proof_like = _proof_like_from_enrollment_transaction(enrollment, tx)
    iid = (tx.get("india_proof_id") or tx.get("india_payment_proof_id") or "").strip()
    if iid:
        proof_doc = await db.india_payment_proofs.find_one({"id": iid}, {"_id": 0})
        if proof_doc and (proof_doc.get("status") or "").strip().lower() == "approved":
            proof_like = _merge_india_proof_into_proof_like(proof_like, proof_doc, tx)
    await _apply_home_coming_emi_payment_sync(
        client_id=pcid,
        proof=proof_like,
        enrollment=enrollment,
        log_prefix="Home Coming EMIs synced (paid enrollment / receipt flow)",
        log_extra_id=tx.get("stripe_session_id") or tx.get("id"),
    )


async def sync_home_coming_emis_after_online_enrollment_payment(session_id: str) -> None:
    """
    After Razorpay verify or Stripe enrollment completion, update ``clients.subscription.emis`` when
    the checkout was Home Coming so the portal schedule matches receipts (GET /student/home).
    """
    sid = (session_id or "").strip()
    if not sid:
        return
    tx = await db.payment_transactions.find_one({"stripe_session_id": sid}, {"_id": 0})
    if not tx:
        return
    txn_clean = {k: v for k, v in tx.items() if k != "_id"}
    await sync_home_coming_emis_for_paid_enrollment_tx(txn_clean)


@router.post("/admin/{proof_id}/approve")
async def approve_payment_proof(proof_id: str):
    """Admin: approve India payment proof and complete enrollment."""
    proof = await db.india_payment_proofs.find_one({"id": proof_id}, {"_id": 0})
    if not proof:
        raise HTTPException(status_code=404, detail="Payment proof not found")

    # Mark proof as approved
    now_approve = datetime.now(timezone.utc).isoformat()
    await db.india_payment_proofs.update_one(
        {"id": proof_id},
        {"$set": {"status": "approved", "approved_at": now_approve, "updated_at": now_approve}},
    )

    enrollment_id = proof.get("enrollment_id")

    program_title_for_match = (proof.get("program_title") or "").strip()
    matched_program = None
    if program_title_for_match:
        matched_program = await db.programs.find_one(
            {"title": {"$regex": program_title_for_match, "$options": "i"}}, {"_id": 0}
        )
        if not matched_program:
            matched_program = await db.programs.find_one(
                {"title": {"$regex": program_title_for_match.split("(")[0].strip(), "$options": "i"}}, {"_id": 0}
            )
    if not matched_program and program_title_for_match:
        matched_program = await db.sessions.find_one(
            {"title": {"$regex": program_title_for_match, "$options": "i"}}, {"_id": 0}
        )

    fallback_item_id = (proof.get("item_id") or "").strip() or (matched_program.get("id") if matched_program else "")
    fallback_item_type = (
        "session"
        if (matched_program and "session_mode" in matched_program and "category" not in matched_program)
        else "program"
    )
    title_for_synthetic = (
        program_title_for_match
        or (proof.get("selected_item") or "").strip()
        or (proof.get("program_type") or "").strip()
    )

    # If no real enrollment exists (manual submissions), create one
    if not enrollment_id or enrollment_id == "MANUAL":
        enrollment_id = f"DIH-{int(datetime.now(timezone.utc).timestamp()) % 100000}-{uuid.uuid4().hex[:3]}"
        new_enrollment = {
            "id": enrollment_id,
            "booker_name": proof.get("payer_name", ""),
            "booker_email": proof.get("booker_email", ""),
            "booker_country": "IN",
            "phone": proof.get("phone", ""),
            "item_type": fallback_item_type,
            "item_id": fallback_item_id,
            "item_title": title_for_synthetic,
            "participant_count": proof.get("participant_count", 1),
            "participants": proof.get("participants", [{"name": proof.get("payer_name", ""), "email": proof.get("booker_email", "")}]),
            "status": "completed",
            "step": 5,
            "payment_method": "manual_proof",
            "bank_name": proof.get("bank_name", ""),
            "is_india_alt": True,
            "paid_at": datetime.now(timezone.utc).isoformat(),
            "created_at": proof.get("created_at", datetime.now(timezone.utc).isoformat()),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.enrollments.insert_one(new_enrollment)
        try:
            await ensure_client_from_enrollment_lead(new_enrollment)
        except Exception as ex:
            logger.warning("ensure_client_from_enrollment_lead after synthetic india enrollment: %s", ex)
        await db.india_payment_proofs.update_one({"id": proof_id}, {"$set": {"enrollment_id": enrollment_id}})

    if enrollment_id:
        full_enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0}) or {}

        resolved_item_id = (
            (full_enrollment.get("item_id") or "").strip()
            or (proof.get("item_id") or "").strip()
            or fallback_item_id
        )
        resolved_item_type = (full_enrollment.get("item_type") or "").strip().lower()
        if resolved_item_type not in ("program", "session"):
            resolved_item_type = fallback_item_type
        resolved_item_title = (
            (full_enrollment.get("item_title") or "").strip()
            or (proof.get("program_title") or "").strip()
            or (proof.get("selected_item") or "").strip()
            or title_for_synthetic
        )
        try:
            resolved_participant_count = int(full_enrollment.get("participant_count") or proof.get("participant_count") or 1)
        except (TypeError, ValueError):
            resolved_participant_count = int(proof.get("participant_count") or 1)
        resolved_participants = full_enrollment.get("participants") or proof.get("participants") or []

        booker_email_resolved = (
            (full_enrollment.get("booker_email") or "").strip()
            or (proof.get("booker_email") or "").strip()
            or (proof.get("payer_email") or "").strip()
        )
        payer_email_resolved = (
            (proof.get("payer_email") or "").strip()
            or (proof.get("booker_email") or "").strip()
            or booker_email_resolved
        )
        booker_name_resolved = (
            (full_enrollment.get("booker_name") or "").strip()
            or (proof.get("booker_name") or "").strip()
            or (proof.get("payer_name") or "").strip()
        )
        phone_resolved = (full_enrollment.get("phone") or "").strip() or (proof.get("payer_phone") or "").strip()

        be_norm = (booker_email_resolved or "").strip().lower()
        pe_norm = (payer_email_resolved or "").strip().lower()
        portal_user_doc = None
        if be_norm:
            portal_user_doc = await db.users.find_one({"email": be_norm}, {"id": 1, "client_id": 1})
        if not portal_user_doc and pe_norm and pe_norm != be_norm:
            portal_user_doc = await db.users.find_one({"email": pe_norm}, {"id": 1, "client_id": 1})

        # Same fields drive receipt email (_send_receipt_and_notifications) and student order history.
        fake_session_id = f"india_{uuid.uuid4().hex[:12]}"
        transaction = {
            "id": str(uuid.uuid4()),
            "enrollment_id": enrollment_id,
            "stripe_session_id": fake_session_id,
            "item_type": resolved_item_type,
            "item_id": resolved_item_id,
            "item_title": resolved_item_title,
            "amount": float(proof.get("amount", 0)),
            "currency": "inr",
            "payment_status": "paid",
            "payment_method": "manual_proof",
            "bank_name": proof.get("bank_name", ""),
            "booker_name": booker_name_resolved,
            "booker_email": be_norm or booker_email_resolved,
            "payer_email": pe_norm or be_norm or payer_email_resolved,
            "phone": phone_resolved,
            "participants": resolved_participants,
            "participant_count": resolved_participant_count,
            "is_india_alt": True,
            "india_proof_id": proof_id,
            "india_payment_method": (proof.get("payment_method") or "").strip().lower(),
            "invoice_number": "",
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc),
        }
        if full_enrollment.get("tier_index") is not None:
            try:
                transaction["tier_index"] = int(full_enrollment["tier_index"])
            except (TypeError, ValueError):
                pass
        if portal_user_doc:
            transaction["portal_user_id"] = portal_user_doc.get("id")
            pcid = (portal_user_doc.get("client_id") or "").strip()
            if pcid:
                transaction["portal_client_id"] = pcid
        # Generate invoice number
        month_prefix = datetime.now(timezone.utc).strftime("%Y-%m")
        count = await db.payment_transactions.count_documents({"invoice_number": {"$regex": f"^{month_prefix}"}})
        transaction["invoice_number"] = f"{month_prefix}-{str(count + 1).zfill(3)}"
        await db.payment_transactions.insert_one(transaction)

        try:
            from routes.points_logic import run_post_payment_loyalty

            txn_clean = {k: v for k, v in transaction.items() if k != "_id"}
            await run_post_payment_loyalty(db, txn_clean)
        except Exception as e:
            logger.warning(f"India proof loyalty points: {e}")

        # Complete enrollment — keep booker contact in sync so receipt email + portal order history match the payer
        enrollment_complete = {
            "step": 5,
            "status": "completed",
            "payment_method": "manual_proof",
            "bank_name": proof.get("bank_name", ""),
            "stripe_session_id": fake_session_id,
            "is_india_alt": True,
            "paid_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        from routes.enrollment_checkout_prepare import snapshot_chosen_tier_from_program

        prog_for_snap = None
        if resolved_item_type == "program" and (resolved_item_id or "").strip():
            pid = (resolved_item_id or "").strip()
            prog_for_snap = await db.programs.find_one({"id": pid}, {"_id": 0, "duration_tiers": 1})
            if not prog_for_snap and pid.isdigit():
                prog_for_snap = await db.programs.find_one({"id": int(pid)}, {"_id": 0, "duration_tiers": 1})
        tix_snap = full_enrollment.get("tier_index")
        if tix_snap is None and proof.get("tier_index") is not None:
            try:
                tix_snap = int(proof["tier_index"])
            except (TypeError, ValueError):
                tix_snap = None
        enrollment_complete.update(snapshot_chosen_tier_from_program(prog_for_snap, tix_snap))
        if (booker_email_resolved or "").strip():
            enrollment_complete["booker_email"] = be_norm or (booker_email_resolved or "").strip().lower()
        if (booker_name_resolved or "").strip():
            enrollment_complete["booker_name"] = (booker_name_resolved or "").strip()
        await db.enrollments.update_one(
            {"id": enrollment_id},
            {"$set": enrollment_complete},
        )

        # Generate UIDs and send receipt email
        try:
            from routes.payments import generate_participant_uids, send_enrollment_receipt
            await generate_participant_uids(fake_session_id)
            # Send receipt directly (not via create_task to avoid silent failures)
            txn_clean = {k: v for k, v in transaction.items() if k != '_id'}
            await send_enrollment_receipt(txn_clean)
            logger.info(
                "Receipt sent for manual proof %s to %s",
                proof_id,
                (booker_email_resolved or proof.get("payer_email") or proof.get("booker_email") or ""),
            )
        except Exception as e:
            logger.warning(f"Error generating UIDs/emails for India payment: {e}")
            import traceback
            traceback.print_exc()

    return {"message": "Payment proof approved. Enrollment completed.", "status": "approved"}


@router.post("/admin/{proof_id}/reject")
async def reject_payment_proof(proof_id: str, reason: str = ""):
    """Admin: reject India payment proof."""
    proof = await db.india_payment_proofs.find_one({"id": proof_id})
    if not proof:
        raise HTTPException(status_code=404, detail="Payment proof not found")

    await db.india_payment_proofs.update_one(
        {"id": proof_id},
        {"$set": {"status": "rejected", "reject_reason": reason, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    enrollment_id = proof.get("enrollment_id")


@router.get("/admin/enrollments")
async def list_enrollments():
    """Admin: list all enrollments with payment details (all checkout paths use `enrollments`)."""
    enrollments = await db.enrollments.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    eids = [e.get("id") for e in enrollments if e.get("id")]
    by_e = await _transactions_grouped_by_enrollment(eids)
    for e in enrollments:
        eid = e.get("id")
        txn = _pick_best_transaction_for_enrollment(e, by_e.get(eid) or [])
        e["payment"] = txn
        e["enrollment_origin"] = _enrollment_origin(e)
        if txn and txn.get("invoice_number"):
            e["invoice_number"] = txn["invoice_number"]
        merged_pp = _merge_participants_for_report(e, txn)
        resolved_title = display_program_title_for_enrollment(e, merged_pp, None)
        if resolved_title:
            e["item_title"] = resolved_title
    return enrollments


@router.get("/admin/enrollments/participant-rows")
async def participant_enrollment_rows(paid_completed_only: bool = False):
    """
    Flat report: one row per participant (never merged into one cell) with enrollment form fields
    including notify, relationship, city/state, attendance, referral, participant email, etc.
    Payment amount/currency is the checkout total (repeated on each row for multi-seat enrollments).
    """
    enrollments = await db.enrollments.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    eids = [e.get("id") for e in enrollments if e.get("id")]
    by_e = await _transactions_grouped_by_enrollment(eids)
    extra_emi_ids, booker_email_client_map = await _booker_email_client_ids_for_emi_overlay(enrollments, by_e)
    base_portal_ids = _client_ids_for_emi_context(enrollments, by_e)
    portal_client_ids = list(dict.fromkeys(list(base_portal_ids) + list(extra_emi_ids)))
    prog_map, cohort_map, cohort_client_map, portal_clients_emi = await asyncio.gather(
        _programs_duration_tiers_by_item_id(enrollments, by_e),
        _portal_cohort_by_contact_emails(enrollments, by_e),
        _portal_cohort_by_client_ids(portal_client_ids),
        _load_portal_clients_emi_context_map(portal_client_ids),
    )
    return build_participant_report_rows(
        enrollments,
        by_e,
        paid_completed_only=paid_completed_only,
        programs_by_id=prog_map,
        portal_cohort_by_email=cohort_map,
        portal_cohort_by_client_id=cohort_client_map,
        portal_clients_by_id=portal_clients_emi,
        booker_email_to_client_id=booker_email_client_map or None,
    )


async def build_participant_report_xlsx_bytes(paid_completed_only: bool = False) -> bytes:
    """Build participant-level enrollment Excel as raw bytes (HTTP download or email attachment)."""
    import io

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    enrollments = await db.enrollments.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    eids = [e.get("id") for e in enrollments if e.get("id")]
    by_e = await _transactions_grouped_by_enrollment(eids)
    extra_emi_ids, booker_email_client_map = await _booker_email_client_ids_for_emi_overlay(enrollments, by_e)
    base_portal_ids = _client_ids_for_emi_context(enrollments, by_e)
    portal_client_ids = list(dict.fromkeys(list(base_portal_ids) + list(extra_emi_ids)))
    prog_map, cohort_map, cohort_client_map, portal_clients_emi = await asyncio.gather(
        _programs_duration_tiers_by_item_id(enrollments, by_e),
        _portal_cohort_by_contact_emails(enrollments, by_e),
        _portal_cohort_by_client_ids(portal_client_ids),
        _load_portal_clients_emi_context_map(portal_client_ids),
    )
    rows = build_participant_report_rows(
        enrollments,
        by_e,
        paid_completed_only=paid_completed_only,
        programs_by_id=prog_map,
        portal_cohort_by_email=cohort_map,
        portal_cohort_by_client_id=cohort_client_map,
        portal_clients_by_id=portal_clients_emi,
        booker_email_to_client_id=booker_email_client_map or None,
    )

    headers = [
        "Invoice #",
        "Enrollment ID",
        "Program",
        "Type",
        "Portal cohort (client)",
        "Program catalog id",
        "Catalog program title",
        "Tier index",
        "Tier label",
        "Chosen start",
        "Chosen end",
        "3-month tier",
        "Enrollment status",
        "Origin",
        "Booker name",
        "Booker email",
        "Booker phone",
        "Booker country",
        "Participant #",
        "Of total",
        "Participant name",
        "Relationship",
        "Age",
        "Gender",
        "Country",
        "City",
        "State",
        "Attendance mode",
        "Notify enrollment",
        "Participant email",
        "Phone",
        "WhatsApp",
        "First time",
        "Referral source",
        "Referred by name",
        "Referred by email",
        "Participant program ID",
        "Participant program title",
        "Participant UID",
        "Payment amount",
        "Currency",
        "Payment status",
        "Created",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    last_col = openpyxl.utils.get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.freeze_panes = "A2"

    for r in rows:
        created = _clean_str(r.get("created_at"))
        if created:
            try:
                from datetime import datetime as dt

                d = dt.fromisoformat(created.replace("Z", "+00:00"))
                created = d.strftime("%Y-%m-%d %H:%M")
            except Exception:
                pass
        ws.append(
            [
                _clean_str(r.get("invoice_number")),
                _clean_str(r.get("enrollment_id")),
                _clean_str(r.get("program")),
                _clean_str(r.get("item_type")),
                _clean_str(r.get("portal_cohort")),
                _clean_str(r.get("program_catalog_id")),
                _clean_str(r.get("catalog_program_title")),
                r.get("tier_index") if r.get("tier_index") is not None else "",
                _clean_str(r.get("tier_label")),
                _clean_str(r.get("chosen_start_date")),
                _clean_str(r.get("chosen_end_date")),
                "yes" if r.get("is_three_month_tier") else "no",
                _clean_str(r.get("enrollment_status")),
                _clean_str(r.get("enrollment_origin")),
                _clean_str(r.get("booker_name")),
                _clean_str(r.get("booker_email")),
                _clean_str(r.get("booker_phone")),
                _clean_str(r.get("booker_country")),
                r.get("participant_index", "") or "",
                r.get("participant_total", "") or "",
                _clean_str(r.get("participant_name")),
                _clean_str(r.get("relationship")),
                _clean_str(r.get("age")),
                _clean_str(r.get("gender")),
                _clean_str(r.get("country")),
                _clean_str(r.get("city")),
                _clean_str(r.get("state")),
                _clean_str(r.get("attendance_mode")),
                _clean_str(r.get("notify_enrollment")),
                _clean_str(r.get("participant_email")),
                _clean_str(r.get("phone")),
                _clean_str(r.get("whatsapp")),
                _clean_str(r.get("is_first_time")),
                _clean_str(r.get("referral_source")),
                _clean_str(r.get("referred_by_name")),
                _clean_str(r.get("referred_by_email")),
                _clean_str(r.get("participant_program_id")),
                _clean_str(r.get("participant_program_title")),
                _clean_str(r.get("participant_uid")),
                r.get("payment_amount", 0) or 0,
                _clean_str(r.get("payment_currency")),
                _clean_str(r.get("payment_status")),
                created,
            ]
        )

    for col in ws.columns:
        max_len = min(max(len(str(cell.value or "")) for cell in col) + 2, 45)
        ws.column_dimensions[col[0].column_letter].width = max_len

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.get("/admin/enrollments/clean-export")
async def export_participant_enrollments_excel(paid_completed_only: bool = False):
    """Excel: one row per participant — clean columns for ops."""
    import io

    try:
        data = await build_participant_report_xlsx_bytes(paid_completed_only=paid_completed_only)
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))

    from fastapi.responses import StreamingResponse

    suffix = "paid_only" if paid_completed_only else "all"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=enrollments_by_participant_{suffix}.xlsx"
        },
    )


@router.get("/admin/enrollments/export")
async def export_enrollments_excel():
    """Admin: export all enrollments as Excel — wide format, one row per enrollment, database-ready."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    def clean(val):
        """Convert any value to a clean string. None/null → empty string."""
        if val is None:
            return ""
        s = str(val)
        if s in ("None", "null"):
            return ""
        return s

    enrollments = await db.enrollments.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    eids = [e.get("id") for e in enrollments if e.get("id")]
    by_e = await _transactions_grouped_by_enrollment(eids)

    # Determine max participant count across all enrollments
    max_participants = 0
    for e in enrollments:
        count = len(e.get("participants") or [])
        if count > max_participants:
            max_participants = count
    max_participants = max(max_participants, 1)

    # Payment data (prefer paid txn when multiple exist)
    for e in enrollments:
        txn = _pick_best_transaction_for_enrollment(e, by_e.get(e.get("id")) or [])
        if txn:
            e["invoice_number"] = txn.get("invoice_number", "")
            e["payment_amount"] = txn.get("amount", 0)
            e["payment_currency"] = txn.get("currency", "")
            e["payment_status_txn"] = txn.get("payment_status", "")
            e["payment_method"] = txn.get("payment_method", "") or e.get("payment_method", "")
            e["bank_name"] = txn.get("bank_name", "") or e.get("bank_name", "")
            e["stripe_session_id"] = txn.get("stripe_session_id", "")

    # Build headers: base columns + per-participant columns
    base_headers = [
        "Invoice #", "Receipt ID", "Status", "Program", "Program Type", "Origin",
        "Booker Name", "Booker Email", "Booker Country", "Booker Phone",
        "Participant Count", "Payment Amount", "Payment Currency", "Payment Method",
        "Bank Account", "Payment Status", "Admin Notes", "Promo Code",
        "VPN Detected", "Enrollment Date",
    ]

    participant_fields = [
        "Name", "Relationship", "Age", "Gender", "Country", "City", "State",
        "Attendance Mode", "Notify", "Is First Time", "Referral Source", "Referred By",
        "Referred By Email", "Email", "Phone", "WhatsApp", "UID",
    ]

    headers = list(base_headers)
    for i in range(1, max_participants + 1):
        for field in participant_fields:
            headers.append(f"Participant {i} {field}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enrollments"

    # Style header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Build column letter ref for auto_filter
    last_col = openpyxl.utils.get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.freeze_panes = "A2"

    for e in enrollments:
        # Format created_at
        created_at = clean(e.get("created_at"))
        if created_at:
            try:
                from datetime import datetime as dt
                d = dt.fromisoformat(created_at.replace("Z", "+00:00")) if isinstance(created_at, str) else created_at
                created_at = d.strftime("%Y-%m-%d %H:%M")
            except Exception:
                pass

        # Base row data
        row = [
            clean(e.get("invoice_number")),
            clean(e.get("id")),
            clean(e.get("status")),
            clean(e.get("item_title")),
            clean(e.get("item_type")),
            _enrollment_origin(e),
            clean(e.get("booker_name")),
            clean(e.get("booker_email")),
            clean(e.get("booker_country")),
            clean(e.get("phone")),
            str(e.get("participant_count", 0) or 0),
            str(e.get("payment_amount", 0) or 0),
            clean(e.get("payment_currency")),
            clean(e.get("payment_method")),
            clean(e.get("bank_name")),
            clean(e.get("payment_status_txn") or e.get("status")),
            clean(e.get("admin_notes")),
            clean(e.get("promo_code")),
            "Yes" if e.get("vpn_detected") else "No",
            created_at,
        ]

        # Append participant columns (wide format)
        participants = e.get("participants") or []
        for i in range(max_participants):
            if i < len(participants):
                p = participants[i]
                p_phone = clean(p.get("phone"))
                p_wa = clean(p.get("whatsapp"))
                notify_yes = _notify_yes_from_participant(p)
                row.extend([
                    clean(p.get("name")),
                    clean(p.get("relationship")),
                    clean(p.get("age")),
                    clean(p.get("gender")),
                    clean(p.get("country")),
                    clean(p.get("city")),
                    clean(p.get("state")),
                    clean(p.get("attendance_mode")),
                    "Yes" if notify_yes else "No",
                    "Yes" if p.get("is_first_time") else "No",
                    clean(p.get("referral_source")),
                    clean(p.get("referred_by_name")),
                    clean(p.get("referred_by_email")),
                    clean(p.get("email")),
                    p_phone,
                    p_wa,
                    clean(p.get("uid")),
                ])
            else:
                # Empty columns for missing participants
                row.extend([""] * len(participant_fields))

        ws.append(row)

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=enrollments.xlsx"}
    )


@router.get("/client-tax/{enrollment_id}")
async def get_client_tax_for_enrollment(enrollment_id: str):
    """Return per-client India tax settings for a given enrollment.

    Used by IndiaPaymentPage to apply the correct tax at checkout.
    Falls back to nulls when no tax is configured for the client.
    """
    enrollment = await db.enrollments.find_one(
        {"id": enrollment_id},
        {"_id": 0, "booker_email": 1, "client_id": 1, "participants": 1, "item_type": 1, "item_id": 1},
    )
    if not enrollment:
        return {"india_tax_enabled": False}

    # Look up client by email or client_id
    client_doc = None
    booker_email = (enrollment.get("booker_email") or "").strip().lower()
    client_id = (enrollment.get("client_id") or "").strip()

    _proj = {
        "_id": 0,
        "india_payment_method": 1,
        "india_discount_percent": 1,
        "india_discount_member_bands": 1,
        "home_coming_india_discount_percent": 1,
        "home_coming_india_discount_member_bands": 1,
        "india_tax_enabled": 1,
        "india_tax_percent": 1,
        "india_tax_label": 1,
        "india_tax_visible_on_dashboard": 1,
    }
    if client_id:
        client_doc = await db.clients.find_one({"id": client_id}, _proj)
    if not client_doc and booker_email:
        client_doc = await db.clients.find_one({"email": booker_email}, _proj)

    if not client_doc:
        return {"india_tax_enabled": False}

    participants = enrollment.get("participants") or []
    participant_count = len(participants) if isinstance(participants, list) else 0
    if participant_count < 1:
        participant_count = 1

    settings_pin = await db.site_settings.find_one(
        {"id": "site_settings"},
        {"_id": 0, "dashboard_sacred_home_annual_program_id": 1},
    )
    pin_hc = str((settings_pin or {}).get("dashboard_sacred_home_annual_program_id") or "").strip()
    chk_ids: List[str] = []
    if enrollment.get("item_type") == "program" and enrollment.get("item_id"):
        chk_ids = [str(enrollment.get("item_id")).strip()]

    from utils.home_coming_discount_scope import filter_client_pricing_for_home_coming_checkout
    from utils.home_coming_crm_fields import home_coming_crm_discount_fields

    _eff_disc = home_coming_crm_discount_fields(client_doc)
    cp_filtered = filter_client_pricing_for_home_coming_checkout(
        {
            "india_discount_percent": _eff_disc["india_discount_percent"],
            "india_discount_member_bands": _eff_disc["india_discount_member_bands"],
        },
        pin_program_id=pin_hc,
        checkout_program_ids=chk_ids,
    )

    return {
        "india_payment_method": client_doc.get("india_payment_method") or None,
        "india_discount_percent": (cp_filtered or {}).get("india_discount_percent"),
        "india_discount_member_bands": (cp_filtered or {}).get("india_discount_member_bands"),
        "participant_count": participant_count,
        "india_tax_enabled": bool(client_doc.get("india_tax_enabled")),
        "india_tax_percent": client_doc.get("india_tax_percent", 18.0),
        "india_tax_label": client_doc.get("india_tax_label", "GST"),
        "india_tax_visible_on_dashboard": client_doc.get("india_tax_visible_on_dashboard", True),
    }
