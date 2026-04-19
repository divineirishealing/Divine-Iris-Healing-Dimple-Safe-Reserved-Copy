"""
Bank / GPay / UPI Transaction Import.
Admin uploads a bank statement or manual log → rows are parsed and stored
in the `bank_transactions` collection → admin tags each row to a client.
No side-effects on enrollments or proofs unless the admin explicitly links them.
"""
from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone
import os
import uuid
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent.parent
load_dotenv(ROOT_DIR / '.env')

router = APIRouter(prefix="/api/bank-transactions", tags=["Bank Transactions"])

_client = AsyncIOMotorClient(os.environ['MONGO_URL'])
db = _client[os.environ['DB_NAME']]
col = db["bank_transactions"]

# ---------------------------------------------------------------------------
# Template column definitions
# ---------------------------------------------------------------------------
TEMPLATE_COLS = [
    ("date",             "Date (DD-MM-YYYY or YYYY-MM-DD)"),
    ("payer_name",       "Payer Name / Description"),
    ("utr_ref",          "UTR / UPI Ref / Transaction ID"),
    ("payment_mode",     "Payment Mode (gpay/upi/bank_transfer/cash_deposit/stripe)"),
    ("amount",           "Amount Received (₹)"),
    ("notes",            "Notes"),
]


# ---------------------------------------------------------------------------
# Download template
# ---------------------------------------------------------------------------
@router.get("/template")
async def download_template():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    sample_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    note_font = Font(italic=True, color="888888", size=9)

    for i, (_, label) in enumerate(TEMPLATE_COLS, 1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    sample = ["15-04-2024", "Priya Sharma", "UTR123456789012", "gpay", "5000", "AWRP April batch"]
    for i, v in enumerate(sample, 1):
        c = ws.cell(row=2, column=i, value=v)
        c.font = note_font
        c.fill = sample_fill

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 30

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bank_transactions_template.xlsx"},
    )


# ---------------------------------------------------------------------------
# Upload & parse
# ---------------------------------------------------------------------------
@router.post("/upload")
async def upload_transactions(file: UploadFile = File(...)):
    """
    Parse an Excel/CSV bank statement or filled template.
    Tries to auto-detect column positions from the header row.
    Stores all rows in bank_transactions with status='untagged'.
    Returns the batch_id so the UI can filter to just this import.
    """
    import io
    from openpyxl import load_workbook

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read file. Upload a valid .xlsx file.")

    ws = wb.active
    header = [str(c.value or "").strip().lower() for c in ws[1]]

    # Map our field names to column index by fuzzy header matching
    ALIASES = {
        "date":         ["date", "transaction date", "txn date", "value date"],
        "payer_name":   ["payer name", "description", "narration", "particulars", "payer", "sender"],
        "utr_ref":      ["utr", "upi ref", "transaction id", "txn id", "ref", "reference", "ref no"],
        "payment_mode": ["payment mode", "mode", "channel", "type"],
        "amount":       ["amount received", "amount", "credit", "inward", "received"],
        "notes":        ["notes", "remarks", "comment"],
    }
    col_idx = {}
    for field, aliases in ALIASES.items():
        for i, h in enumerate(header):
            if any(a in h for a in aliases):
                col_idx[field] = i
                break

    def get(row, field):
        i = col_idx.get(field)
        v = row[i] if i is not None and i < len(row) else None
        return str(v).strip() if v is not None and v != "" else ""

    batch_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        amount_raw = get(row, "amount").replace(",", "").replace("₹", "").strip()
        try:
            amount = float(amount_raw) if amount_raw else None
        except ValueError:
            amount = None

        if amount is None or amount <= 0:
            skipped += 1
            continue

        doc = {
            "id":           str(uuid.uuid4()),
            "batch_id":     batch_id,
            "date":         get(row, "date"),
            "payer_name":   get(row, "payer_name"),
            "utr_ref":      get(row, "utr_ref"),
            "payment_mode": get(row, "payment_mode").lower() or "upi",
            "amount":       amount,
            "notes":        get(row, "notes"),
            "status":       "untagged",
            "client_id":    None,
            "client_email": None,
            "client_name":  None,
            "created_at":   now,
        }
        await col.insert_one(doc)
        inserted += 1

    return {"batch_id": batch_id, "inserted": inserted, "skipped": skipped}


# ---------------------------------------------------------------------------
# List transactions
# ---------------------------------------------------------------------------
@router.get("")
async def list_transactions(status: Optional[str] = None, batch_id: Optional[str] = None):
    query = {}
    if status:
        query["status"] = status
    if batch_id:
        query["batch_id"] = batch_id
    docs = await col.find(query, {"_id": 0}).sort("date", -1).to_list(2000)
    return docs


# ---------------------------------------------------------------------------
# Tag a transaction to a client
# ---------------------------------------------------------------------------
class TagRequest(BaseModel):
    client_id: str
    client_email: Optional[str] = None
    client_name: Optional[str] = None


@router.put("/{txn_id}/tag")
async def tag_transaction(txn_id: str, body: TagRequest):
    txn = await col.find_one({"id": txn_id})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")

    # Optionally back-fill payment method on the client record
    client_doc = await db.clients.find_one({"id": body.client_id}, {"_id": 0, "india_payment_method": 1})
    if client_doc and not client_doc.get("india_payment_method"):
        pm = (txn.get("payment_mode") or "").strip().lower()
        if pm:
            await db.clients.update_one(
                {"id": body.client_id},
                {"$set": {"india_payment_method": pm, "updated_at": datetime.now(timezone.utc).isoformat()}},
            )

    await col.update_one(
        {"id": txn_id},
        {"$set": {
            "status":       "tagged",
            "client_id":    body.client_id,
            "client_email": body.client_email,
            "client_name":  body.client_name,
            "tagged_at":    datetime.now(timezone.utc).isoformat(),
        }},
    )
    return {"message": "Tagged"}


@router.put("/{txn_id}/untag")
async def untag_transaction(txn_id: str):
    txn = await col.find_one({"id": txn_id})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    await col.update_one(
        {"id": txn_id},
        {"$set": {"status": "untagged", "client_id": None, "client_email": None, "client_name": None}},
    )
    return {"message": "Untagged"}


@router.delete("/{txn_id}")
async def delete_transaction(txn_id: str):
    await col.delete_one({"id": txn_id})
    return {"message": "Deleted"}
